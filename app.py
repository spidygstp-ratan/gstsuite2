# app.py — GST Reconciliation Tool Enterprise v9.0
# ══════════════════════════════════════════════════════
# LICENSE GATE — runs before any app logic
# ══════════════════════════════════════════════════════
import streamlit as _st

# We need page config first, so set it here early if not already set
# (The real set_page_config is called again below — Streamlit ignores duplicates)
try:
    from modules.license_manager import get_license_status, activate_key, is_allowed_to_run

    _lic = get_license_status()

    # ── Blocked (wrong device) or expired — show gate and STOP ────────────────
    if _lic["status"] in ("expired_trial", "expired_key", "blocked"):
        _st.set_page_config(page_title="GST Tool — Activation Required", page_icon="🔐", layout="centered")
        _st.markdown("""
        <style>
        .stApp { background: linear-gradient(135deg, #1e3a5f 0%, #0f52ba 100%); }
        .act-card {
            background: white; border-radius: 20px; padding: 40px 50px;
            max-width: 500px; margin: 60px auto; box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            text-align: center;
        }
        </style>""", unsafe_allow_html=True)

        _st.markdown("""
        <div class="act-card">
            <h1>🔐 GST Reconciliation Tool</h1>
            <h3 style="color:#e74c3c;">Activation Required</h3>
        </div>""", unsafe_allow_html=True)

        _st.error(f"🚫 {_lic['message']}")
        _st.info(f"📟 Your Device ID: `{_lic['mac']}`")
        _st.markdown("---")
        _st.markdown("### Enter Your Activation Key")
        _st.caption("Purchase a key from the seller and enter it below.")

        _key_input = _st.text_input("Activation Key", placeholder="XXXX-XXXX-XXXX-XXXX", max_chars=19)
        if _st.button("🔓 Activate Software", type="primary", use_container_width=True):
            if _key_input:
                _result = activate_key(_key_input)
                if _result["success"]:
                    _st.success(_result["message"])
                    _st.balloons()
                    _st.rerun()
                else:
                    _st.error(_result["message"])
            else:
                _st.warning("Please enter your activation key first.")
        _st.stop()  # ← Hard stop. App code below does NOT run.

    # ── Trial or Active — show small banner and continue ──────────────────────
    elif _lic["status"] == "trial":
        _trial_days = _lic["days_left"]
        # Store banner info in session for display after page config
        import streamlit as _st2
        if "lic_banner" not in _st2.session_state:
            _st2.session_state["lic_banner"] = ("trial", _lic["message"])

    elif _lic["status"] == "active":
        if "lic_banner" not in _st.session_state:
            _st.session_state["lic_banner"] = ("active", _lic["message"])

except Exception as _lic_err:
    pass  # If license module fails, allow run (fail open for safety)
# ══════════════════════════════════════════════════════
# END LICENSE GATE
# ══════════════════════════════════════════════════════
# CHANGES FROM v3.2:
#   BUG FIXES:
#     1. Unique_ID assigned before process_dataset() — manual links stable across re-runs
#     2. failed_matches no longer double-counted (core_engine fix)
#     3. Tab 2 Suggestions filter fixed (was if/elif logic bug)
#     4. gst_scraper import wrapped in try/except — no crash if file missing
#     5. CDNR results now saved to DB and restored when loading from History
#     6. K4 numeric key now has 4-digit minimum + tolerance check (false-match prevention)
#   FEATURES ADDED:
#     7. GSTIN format validator in setup stage
#     8. Template download for standard column mapping
#     9. Per-vendor tolerance override table
#    10. Combined B2B + CDNR Net ITC banner on Dashboard
#    11. Data confidence panel (Books vs GSTR-2B row/value summary) after upload
#    12. Executive Summary in Excel report (v4 report_gen)
#    13. CDNR Suggestions sub-view showing GSTIN match status
#    14. ITC Impact section removed from CDNR tab
#    15. Audit log for manual actions
#    16. Streamlit-lottie removed — uses native progress bar

import streamlit as st
import pandas as pd
import numpy as np
import altair as alt
import time
import urllib.parse
import io
import zipfile
import os
import re

# --- CORE IMPORTS ---
from modules.constants import (REQUIRED_FIELDS, FIXED_BOOKS_MAPPING, FIXED_GST_MAPPING,
                               SOFTWARE_COLUMN_PROFILES)
from modules.data_utils     import (load_data_preview, find_best_match,
                                    extract_meta_from_readme, standardize_invoice_numbers)
from modules.core_engine    import run_reconciliation
from modules.report_gen     import generate_excel, generate_vendor_split_zip
from modules.utils          import show_processing_animation
from modules.email_tool     import (get_vendors_with_issues, generate_email_draft,
                                    generate_whatsapp_message, generate_whatsapp_message_multilang,
                                    generate_targeted_notice, get_vendors_by_category)
from modules.notice_importer import parse_uploaded_result_excel, get_available_sheets
from modules.pdf_gen        import create_vendor_pdf, create_itc_risk_pdf, create_action_report_pdf
from modules.db_handler     import (init_db, save_reconciliation, get_history_list,
                                    load_reconciliation, delete_reconciliation,
                                    save_cdnr_to_history, log_action, get_audit_log,
                                    upsert_followup, get_followups, update_followup_status,
                                    save_followup_notice_sent, get_overdue_followups,
                                    get_all_clients_itc_summary, compare_two_recons)
from modules.file_manager   import get_client_path, save_file_to_folder, open_folder

# --- PRE-PROCESSORS ---
from modules.pre_processor  import smart_read_b2ba, process_amendments

# --- CDNR ENGINE ---
from modules.cdnr_processor    import process_cdnr_reconciliation
from modules.cdnr_report_gen   import generate_cdnr_excel
from modules.combined_report_gen import generate_combined_excel

# --- MAIN DASHBOARD ---
from modules.dashboard_ui import render_dashboard

# ==========================================
# PAGE CONFIG & CSS
# ==========================================
st.set_page_config(
    page_title="GST Reconciliation Tool",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded"
)
init_db()

st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');

    /* ── ROOT VARIABLES ── */
    :root {
        --navy: #0D1B40; --blue: #1352C9; --blue-mid: #2563EB;
        --blue-light: #EEF4FF; --red: #C7000A; --red-light: #FFF1F2;
        --red-lt: #FFF1F2;
        --gold: #B45309; --gold-light: #FFFBEB;
        --green: #0F6B3C; --green-light: #F0FDF7; --green-lt: #F0FDF7;
        --gray-50: #F8FAFC; --gray-100: #F1F5F9;
        --gray-200: #E2E8F0; --gray-400: #94A3B8; --gray-600: #475569;
        --gray-800: #1E293B; --white: #FFFFFF;
        --bg: #EEF2F8; --cream: #F8FAFC;
        --amber: #1352C9; --amber-lt: #EEF4FF; --amber-dk: #0D1B40; --amber-md: #2563EB;
        --navy2: #1352C9; --text: #0D1B40; --t2: #475569; --t3: #94A3B8;
        --border: #E2E8F0;
        --shadow: 0 1px 3px rgba(0,0,0,0.06); --shadow-lg: 0 4px 20px rgba(13,27,64,0.12);
        --r: 14px; --r-sm: 10px; --r-xs: 8px;
    }

    /* ── APP BACKGROUND ── */
    .stApp { background-color: #EEF2F8 !important; font-family: 'DM Sans', sans-serif !important; font-size: 14px !important; }

    /* ── NEXT STEP BLINK ANIMATION ── */
    @keyframes gst-pulse {
        0%   { box-shadow: 0 0 0 0 rgba(19,82,201,0.55); border-color: #1352C9; }
        60%  { box-shadow: 0 0 0 8px rgba(19,82,201,0); border-color: #1352C9; }
        100% { box-shadow: 0 0 0 0 rgba(19,82,201,0); border-color: #1352C9; }
    }
    @keyframes gst-arrow-bounce {
        0%, 100% { transform: translateX(0); }
        50%       { transform: translateX(5px); }
    }
    .next-step-hint {
        border: 2px solid #1352C9 !important;
        border-radius: 12px !important;
        padding: 14px 18px !important;
        background: linear-gradient(135deg, #EEF4FF 0%, #DBEAFE 100%) !important;
        animation: gst-pulse 1.6s ease-in-out infinite !important;
        margin-top: 12px !important;
    }
    .next-step-arrow {
        display: inline-block;
        animation: gst-arrow-bounce 0.9s ease-in-out infinite;
    }

    /* ── STRONGER TYPOGRAPHY ── */
    h1, h2, h3 { font-weight: 800 !important; letter-spacing: -0.025em !important; color: #0D1B40 !important; }
    .stMarkdown p { font-size: 13px !important; line-height: 1.6 !important; }
    label, .stSelectbox label, .stTextInput label { font-weight: 700 !important; font-size: 12px !important; color: #1E293B !important; }
    .stTabs [data-baseweb="tab"] { font-weight: 700 !important; font-size: 12px !important; letter-spacing: -0.01em !important; }
    .stTabs [data-baseweb="tab"][aria-selected="true"] { font-weight: 800 !important; font-size: 12px !important; }
    div[data-testid="stMetricValue"] { font-weight: 800 !important; }
    div[data-testid="stMetricLabel"] { font-weight: 700 !important; }
    .stCaption, .stCaption p { font-size: 12px !important; color: #475569 !important; font-weight: 500 !important; }
    .stSuccess p, .stInfo p, .stWarning p, .stError p { font-weight: 600 !important; font-size: 13px !important; }
    .main .block-container { padding-top: 1.2rem !important; padding-bottom: 2rem !important; }

    /* ── NO WHITE FLASH ── */
    @keyframes gst-fadein { from{opacity:0;transform:translateY(10px)} to{opacity:1;transform:translateY(0)} }
    .main .block-container { animation: gst-fadein 0.3s ease both !important; }
    html, body, [data-testid="stAppViewContainer"] { background-color: var(--bg) !important; }

    /* ── SIDEBAR ── */
    [data-testid="stSidebar"] > div:first-child { background: var(--navy) !important; border-right: none !important; }
    [data-testid="stSidebar"] * { color: rgba(255,255,255,0.85) !important; }
    [data-testid="stSidebar"] .stTextInput input {
        background: rgba(255,255,255,0.07) !important; border: 1px solid rgba(255,255,255,0.12) !important;
        border-radius: 8px !important; color: rgba(255,255,255,0.7) !important; font-size: 12px !important;
    }
    [data-testid="stSidebar"] .stButton > button {
        background: rgba(255,255,255,0.07) !important; border: 1px solid rgba(255,255,255,0.12) !important;
        color: rgba(255,255,255,0.75) !important; border-radius: 50px !important; font-size: 12px !important;
    }
    [data-testid="stSidebar"] .stButton > button:hover { background: rgba(37,99,235,0.25) !important; border-color: rgba(37,99,235,0.5) !important; }

    /* ── METRIC CARDS — amber accent ── */
    div[data-testid="stMetric"] {
        background: var(--white) !important; border: 1.5px solid #E8E5E0 !important;
        border-radius: var(--r-sm) !important; padding: 16px 18px !important;
        box-shadow: var(--shadow) !important; transition: all 0.2s ease !important;
        border-top: 3px solid var(--amber) !important;
    }
    div[data-testid="stMetric"]:hover { transform: translateY(-2px) !important; box-shadow: var(--shadow-lg) !important; }
    div[data-testid="stMetricLabel"] {
        font-size: 10px !important; color: var(--t3) !important; font-weight: 800 !important;
        text-transform: uppercase !important; letter-spacing: 0.07em !important;
        font-family: 'Plus Jakarta Sans', sans-serif !important;
    }
    div[data-testid="stMetricValue"] {
        font-size: 22px !important; color: var(--navy) !important; font-weight: 800 !important;
        letter-spacing: -0.025em !important; font-family: 'Plus Jakarta Sans', sans-serif !important;
    }
    div[data-testid="stMetricDelta"] { font-size: 11px !important; }

    /* ── TABS — navy active ── */
    .stTabs [data-baseweb="tab-list"] {
        gap: 3px !important; background: var(--white) !important;
        border-radius: var(--r-xs) !important; padding: 5px !important;
        box-shadow: var(--shadow) !important; border: none !important; margin-bottom: 16px !important;
    }
    .stTabs [data-baseweb="tab"] {
        height: 36px !important; background: transparent !important;
        border-radius: var(--r-xs) !important; padding: 0 14px !important;
        font-weight: 700 !important; color: var(--t2) !important;
        border: none !important; font-family: 'Plus Jakarta Sans', sans-serif !important; font-size: 11px !important;
    }
    .stTabs [data-baseweb="tab"]:hover { background: #F0EDE8 !important; color: var(--text) !important; }
    .stTabs [data-baseweb="tab"][aria-selected="true"] {
        background: var(--navy) !important; color: #fff !important;
        box-shadow: 0 2px 8px rgba(13,27,64,0.25) !important;
    }

    /* ── BUTTONS ── */
    div.stButton > button {
        border-radius: 50px !important; font-weight: 700 !important;
        font-family: 'Plus Jakarta Sans', sans-serif !important;
        padding: 0.4rem 1.1rem !important; transition: all 0.18s !important; font-size: 12px !important;
    }
    div.stButton > button[kind="primary"] {
        background: var(--navy) !important; color: #fff !important; border: none !important;
        box-shadow: 0 4px 12px rgba(27,32,53,.25) !important;
    }
    div.stButton > button[kind="primary"]:hover { background: var(--navy2) !important; transform: translateY(-1px) !important; }

    /* ── INPUTS ── */
    .stTextInput input, .stNumberInput input, .stSelectbox select {
        border-radius: var(--r-xs) !important; border: 1.5px solid #E0DDD8 !important;
        background: var(--cream) !important; font-family: 'Plus Jakarta Sans', sans-serif !important;
        font-size: 12px !important; transition: border-color 0.18s !important;
    }
    .stTextInput input:focus, .stNumberInput input:focus { border-color: var(--navy) !important; box-shadow: 0 0 0 3px rgba(27,32,53,.08) !important; }

    /* ── EXPANDERS ── */
    .streamlit-expanderHeader {
        background: var(--white) !important; border-radius: var(--r-xs) !important;
        border: 1.5px solid #E0DDD8 !important; font-weight: 700 !important; font-size: 12px !important;
    }
    .streamlit-expanderContent { border: 1.5px solid #E0DDD8 !important; border-top: none !important; border-radius: 0 0 var(--r-xs) var(--r-xs) !important; }

    /* ── DATAFRAME ── */
    [data-testid="stDataFrame"] { border-radius: var(--r-xs) !important; overflow: hidden !important; }

    /* ── DOWNLOAD BUTTON ── */
    div.stDownloadButton > button {
        border-radius: 50px !important; font-weight: 700 !important;
        font-family: 'Plus Jakarta Sans', sans-serif !important; font-size: 12px !important;
    }
    div.stDownloadButton > button[kind="primary"] {
        background: var(--navy) !important; border: none !important; color: #fff !important;
        box-shadow: 0 4px 12px rgba(27,32,53,.25) !important;
    }

    /* ── ALERTS ── */
    .stSuccess, .stInfo, .stWarning, .stError { border-radius: var(--r-xs) !important; font-size: 12px !important; }

    /* ════ LEGACY CUSTOM CLASSES (still used in non-dashboard pages) ════ */

    /* App header / success bar */
    .app-header-banner {
        background: var(--navy); padding: 16px 24px; border-radius: var(--r-sm); margin-bottom: 16px;
        display: flex; align-items: center; justify-content: space-between;
        box-shadow: var(--shadow-lg);
    }
    .header-title-text { font-size: 18px; font-weight: 800; color: #fff; letter-spacing: -0.025em; }
    .header-sub-text   { font-size: 11px; color: rgba(255,255,255,0.45); margin-top: 2px; }
    .header-version-pill {
        background: rgba(255,255,255,0.12); color: rgba(255,255,255,0.85);
        font-size: 9px; font-weight: 800; padding: 3px 10px; border-radius: 50px;
        border: 1px solid rgba(255,255,255,0.2); margin-left: 10px; vertical-align: middle;
    }
    .header-client-pill {
        background: var(--amber); border-radius: 50px; padding: 5px 14px;
        font-size: 11px; font-weight: 800; color: var(--navy); display: inline-block;
    }

    /* Sidebar history cards */
    .hist-card { background: rgba(255,255,255,.05); border: 1px solid rgba(255,255,255,.08); border-radius: 10px; padding: 10px 12px; margin-bottom: 6px; }
    .hist-card-name { font-size: 12px; font-weight: 700; color: #fff; margin-bottom: 3px; }
    .hist-card-meta { font-size: 10px; color: rgba(255,255,255,.4); display: flex; align-items: center; gap: 5px; }
    .dot { width: 7px; height: 7px; border-radius: 50%; display: inline-block; flex-shrink: 0; }
    .dot-red   { background: #F87171; }
    .dot-green { background: var(--amber); }
    .dot-gold  { background: var(--amber); }

    /* Step wizard */
    .step-wizard { display: flex; align-items: center; gap: 0; margin-bottom: 20px; }
    .step-node   { display: flex; align-items: center; gap: 8px; }
    .step-circle { width: 28px; height: 28px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 11px; font-weight: 800; flex-shrink: 0; }
    .step-circle-done   { background: var(--green); color: #fff; }
    .step-circle-active { background: var(--navy); color: #fff; box-shadow: 0 0 0 4px rgba(27,32,53,.12); }
    .step-circle-idle   { background: #EDEAE4; color: var(--t3); }
    .step-text-active { font-size: 12px; font-weight: 700; color: var(--navy); }
    .step-text-done   { font-size: 12px; font-weight: 700; color: var(--green); }
    .step-text-idle   { font-size: 12px; font-weight: 600; color: var(--t3); }
    .step-line      { flex: 1; height: 2px; background: #EDEAE4; margin: 0 10px; min-width: 30px; }
    .step-line-done { background: var(--green); }

    /* Software cards */
    .sw-grid { display: grid; grid-template-columns: repeat(4,1fr); gap: 10px; margin-bottom: 20px; }
    .sw-card { background: var(--cream); border: 2px solid transparent; border-radius: var(--r-xs); padding: 14px 10px; cursor: pointer; transition: all 0.18s; text-align: center; }
    .sw-card:hover { border-color: #D0CECA; }
    .sw-card-selected { border-color: var(--amber) !important; background: var(--amber-lt) !important; }
    .sw-icon { font-size: 20px; margin-bottom: 5px; }
    .sw-name { font-size: 10px; font-weight: 700; color: var(--t2); }

    /* Upload cards */
    .upload-card { border: 2px dashed #D0CECA; border-radius: var(--r-sm); padding: 26px 20px; text-align: center; background: var(--cream); transition: all 0.2s; }
    .upload-card:hover { border-color: var(--navy); }
    .upload-card-filled { border-style: solid !important; border-color: var(--green) !important; background: var(--green-lt) !important; }

    /* ITC banner (inside results Tab 1) */
    .itc-net-banner {
        background: var(--white); border-radius: var(--r); padding: 16px 22px;
        display: flex; align-items: center; box-shadow: var(--shadow); margin-bottom: 16px;
    }
    .itc-item { flex: 1; text-align: center; }
    .itc-item-label { font-size: 9px; font-weight: 800; text-transform: uppercase; letter-spacing: .07em; color: var(--t3); margin-bottom: 4px; }
    .itc-item-val { font-size: 18px; font-weight: 800; color: var(--navy); }
    .itc-item-val-red  { color: var(--red) !important; }
    .itc-item-val-blue { color: #1D49B5 !important; }
    .itc-item-val-grn  { color: var(--green) !important; }
    .itc-sep { width: 1px; height: 36px; background: #EDEAE4; }

    /* Result KPI tiles */
    .rk { background: var(--white); border-radius: var(--r-sm); padding: 15px 16px; box-shadow: var(--shadow); }
    .rk.red { background: var(--red-lt); }
    .rk.green { background: var(--green-lt); }
    .rk.amber { background: var(--amber-lt); }
    .rk-l { font-size: 9px; font-weight: 800; text-transform: uppercase; letter-spacing: .07em; color: var(--t3); margin-bottom: 6px; }
    .rk.red .rk-l   { color: #A82C18; }
    .rk.green .rk-l { color: #0E5C37; }
    .rk.amber .rk-l { color: var(--amber-md); }
    .rk-v { font-size: 22px; font-weight: 800; color: var(--navy); letter-spacing: -.025em; }
    .rk.red .rk-v   { color: var(--red); }
    .rk.green .rk-v { color: var(--green); }
    .rk.amber .rk-v { color: var(--amber-dk); }

    /* Progress bar */
    .pbar { height: 5px; background: rgba(0,0,0,.06); border-radius: 50px; margin-top: 10px; overflow: hidden; }
    .pb-f { height: 100%; border-radius: 50px; }

    /* Risk radar rows */
    .risk-row { display: flex; align-items: center; gap: 10px; padding: 10px 0; border-bottom: 1px solid #F5F2ED; }
    .risk-row:last-child { border: none; }
    .risk-rank { width: 22px; height: 22px; border-radius: 6px; background: var(--red-lt); color: var(--red); font-size: 10px; font-weight: 800; display: grid; place-items: center; flex-shrink: 0; }
    .risk-name { font-size: 12px; font-weight: 700; color: var(--navy); }
    .risk-sub  { font-size: 10px; color: var(--t3); margin-top: 1px; }
    .risk-amount { font-size: 12px; font-weight: 800; color: var(--red); white-space: nowrap; }

    /* Download dark card */
    .dl-card { background: var(--navy); border-radius: var(--r); padding: 22px; text-align: center; margin-top: 14px; box-shadow: var(--shadow-lg); }
    .dl-btn { background: var(--amber); color: var(--navy); font-weight: 800; border: none; border-radius: 50px; padding: 10px 24px; font-size: 13px; cursor: pointer; font-family: 'Plus Jakarta Sans', sans-serif; }

    /* Success bar */
    .recon-success-bar {
        background: var(--green-lt); border-radius: var(--r-sm); padding: 12px 18px; margin-bottom: 16px;
        display: flex; align-items: center; justify-content: space-between;
        box-shadow: 0 2px 12px rgba(24,137,90,.08);
    }
    .recon-success-text { font-size: 12px; font-weight: 700; color: #0E5C37; }
    .recon-success-dot { width: 8px; height: 8px; background: var(--green); border-radius: 50%; margin-right: 9px; }

    /* Confidence boxes */
    .confidence-box { background: var(--amber-lt); border-radius: var(--r-xs); padding: 13px 15px; margin: 6px 0; }
    .confidence-box-gst { background: #EEF4FF; border-radius: var(--r-xs); padding: 13px 15px; margin: 6px 0; }
    .confidence-title     { font-weight: 800; color: var(--navy); font-size: 12px; margin-bottom: 5px; }
    .confidence-title-gst { font-weight: 800; color: var(--navy); font-size: 12px; margin-bottom: 5px; }

    /* Period chips */
    .month-chip-row { display: flex; flex-wrap: wrap; gap: 6px; margin-top: 6px; }
    .chip-sel   { padding: 4px 12px; border-radius: 50px; font-size: 11px; font-weight: 700; background: var(--amber); color: var(--navy); border: none; cursor: pointer; }
    .chip-unsel { padding: 4px 12px; border-radius: 50px; font-size: 11px; font-weight: 600; background: var(--cream); color: var(--t2); border: 1.5px solid #D0CECA; cursor: pointer; }

    /* Section headers */
    .section-hdr { font-size: 14px; font-weight: 800; color: var(--navy); margin: 0 0 3px 0; }
    .section-sub { font-size: 11px; color: var(--t3); margin-bottom: 14px; }

    /* Overdue alert sidebar */
    .overdue-alert { background: rgba(214,57,32,.18); border-radius: 10px; padding: 10px 12px; margin: 6px 0; }
    .overdue-text { font-size: 11px; font-weight: 700; color: #FCA5A5 !important; }
    .overdue-sub  { font-size: 10px; color: rgba(255,255,255,.3) !important; margin-top: 2px; }

    /* Dash sidebar btn */
    .dash-side-btn { background: rgba(242,197,33,.15); border: 1px solid rgba(242,197,33,.25); border-radius: var(--r-xs); padding: 10px 12px; cursor: pointer; margin: 4px 0; }
    .dash-side-btn-text { font-size: 12px; font-weight: 700; color: var(--amber) !important; }
    .dash-side-btn-sub  { font-size: 10px; color: rgba(255,255,255,.35) !important; margin-top: 2px; }
    </style>
""", unsafe_allow_html=True)

# ==========================================
# SESSION STATE INIT
# ==========================================
defaults = {
    'show_dashboard':      True,
    'app_stage':           'setup',
    'manual_matches':      [],
    'current_client_path': None,
    'cdnr_result':         None,
    'combined_report_bytes': None,
    'cdnr_summary':        None,
    'current_recon_id':    None,
    'vendor_tolerances':   {},
    'data_summary_books':  None,
    'data_summary_gst':    None,
    'imp_wa_preview':      None,
    'wa_lang':             'en',   # WhatsApp language: 'en','hi','gu'
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ==========================================
# HELPERS
# ==========================================
def save_callback(folder_path, file_name, file_data):
    if folder_path:
        save_file_to_folder(folder_path, file_name, file_data)

def validate_gstin(gstin: str) -> bool:
    pattern = r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$'
    return bool(re.match(pattern, str(gstin).strip().upper()))

def make_data_summary(df, label):
    """Build quick stats dict from a raw DataFrame — including GST components."""
    try:
        gstin_col = next((c for c in df.columns if 'gstin' in c.lower()), None)
        tax_col   = next((c for c in df.columns if 'taxable' in c.lower()), None)
        inv_col   = next((c for c in df.columns if 'invoice' in c.lower() and 'number' in c.lower()), None)
        igst_col  = next((c for c in df.columns if 'integrated' in c.lower() or c.lower() == 'igst'), None)
        cgst_col  = next((c for c in df.columns if 'central' in c.lower() or c.lower() == 'cgst'), None)
        sgst_col  = next((c for c in df.columns if 'state' in c.lower() or c.lower() in ('sgst','sgst/utgst')), None)
        def _sum(col):
            if col:
                return pd.to_numeric(df[col].astype(str).str.replace(',',''), errors='coerce').sum()
            return 0.0
        n_rows    = len(df)
        n_gstin   = df[gstin_col].nunique() if gstin_col else '—'
        total_tax = _sum(tax_col)
        n_inv     = df[inv_col].nunique() if inv_col else n_rows
        total_igst = _sum(igst_col)
        total_cgst = _sum(cgst_col)
        total_sgst = _sum(sgst_col)
        total_gst  = total_igst + total_cgst + total_sgst
        return {'label': label, 'rows': n_rows, 'invoices': n_inv, 'gstins': n_gstin,
                'taxable': total_tax, 'igst': total_igst, 'cgst': total_cgst,
                'sgst': total_sgst, 'total_gst': total_gst}
    except Exception:
        return {'label': label, 'rows': '?', 'invoices': '?', 'gstins': '?',
                'taxable': 0, 'igst': 0, 'cgst': 0, 'sgst': 0, 'total_gst': 0}

def make_template_excel():
    """Generates a blank Purchase Register template Excel."""
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine='xlsxwriter') as writer:
        headers = list(FIXED_BOOKS_MAPPING.values())
        headers = [h for h in headers if h != '<No Column / Blank>']
        sample  = {
            'GSTIN of Supplier': '24ABCDE1234F1Z5',
            'Invoice Number':     'INV-001',
            'Invoice date':       '01/04/2025',
            'Invoice Value':      '11800',
            'Taxable Value':      '10000',
            'Integrated Tax Paid': '1800',
            'Central Tax Paid':   '',
            'State/UT Tax Paid':  '',
            'Cess Paid':          '',
            'Place Of Supply':    '24-Gujarat',
            'Reverse Charge':     'N',
        }
        df_tmpl = pd.DataFrame([sample])
        df_tmpl.to_excel(writer, sheet_name='Purchase Register', index=False)
        wb  = writer.book
        ws  = writer.sheets['Purchase Register']
        hdr = wb.add_format({'bold': True, 'bg_color': '#4472C4', 'font_color': 'white', 'border': 1})
        for i, col in enumerate(df_tmpl.columns):
            ws.write(0, i, col, hdr)
            ws.set_column(i, i, 22)
    out.seek(0)
    return out.getvalue()

# ==========================================
# SIDEBAR — HISTORY
# ==========================================
with st.sidebar:
    # ── Back to Dashboard ─────────────────────────────────────────────────────
    if not st.session_state.get('show_dashboard', True):
        if st.button("🏠  ← Dashboard", key="back_to_dashboard", use_container_width=True, type="primary"):
            st.session_state['show_dashboard'] = True
            st.rerun()
        st.markdown("<hr style='border-color:rgba(255,255,255,0.08);margin:8px 0 12px;'>", unsafe_allow_html=True)

    st.markdown("""
    <div style="padding:18px 4px 12px;">
        <div style="font-size:15px;font-weight:700;color:#fff;">🛡️ GST Recon Tool</div>
        <div style="font-size:11px;color:rgba(255,255,255,0.4);margin-top:2px;">Enterprise Edition</div>
        <div style="display:inline-block;background:rgba(37,99,235,0.35);color:#93C5FD;
                    font-size:10px;font-weight:700;padding:2px 9px;border-radius:20px;
                    margin-top:6px;letter-spacing:0.5px;">v7.0</div>
    </div>
    <hr style="border-color:rgba(255,255,255,0.08);margin:0 0 10px;">
    """, unsafe_allow_html=True)

    st.markdown("<div style='font-size:10px;font-weight:700;color:rgba(255,255,255,0.35);text-transform:uppercase;letter-spacing:1px;margin-bottom:6px;'>Client History</div>", unsafe_allow_html=True)
    search_query = st.text_input("Search client", placeholder="🔍  Search client...", label_visibility="collapsed").lower()
    history_df   = get_history_list()

    if not history_df.empty:
        if search_query:
            history_df = history_df[
                history_df['company_name'].str.lower().str.contains(search_query, na=False) |
                history_df['gstin'].str.lower().str.contains(search_query, na=False)
            ]
        if history_df.empty:
            st.warning("No matching records found.")
        else:
            for idx, row in history_df.iterrows():
                # Determine status dot color based on period/data
                dot_color = "dot-gold"
                with st.container():
                    st.markdown(f"""
                    <div class="hist-card">
                        <div class="hist-card-name">{row['company_name']}</div>
                        <div class="hist-card-meta">
                            <span class="dot {dot_color}"></span>
                            {row.get('fy','—')} &middot; {row['period']} &middot; {str(row['timestamp'])[:10]}
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                    c_open, c_del = st.columns([3, 1])
                    with c_open:
                        if st.button("📂 Open", key=f"hist_{row['id']}", use_container_width=True):
                            meta, df_loaded, df_cdnr, cdnr_summary = load_reconciliation(row['id'])
                            st.session_state['last_result']      = df_loaded
                            st.session_state['df_b_clean']       = df_loaded
                            st.session_state['df_g_clean']       = df_loaded
                            st.session_state['meta_gstin']       = meta['gstin']
                            st.session_state['meta_name']        = meta['company_name']
                            st.session_state['meta_fy']          = meta['fy']
                            st.session_state['meta_period']      = meta['period']
                            st.session_state.current_client_path = get_client_path(
                                meta['company_name'], meta['gstin'], meta['fy'], meta['period'])
                            st.session_state.current_recon_id    = row['id']
                            st.session_state.cdnr_result  = df_cdnr
                            st.session_state.cdnr_summary = cdnr_summary
                            st.session_state['file_books_bytes'] = None
                            st.session_state['file_gst_bytes']   = None
                            st.session_state.app_stage = 'results'
                            st.rerun()
                    with c_del:
                        if st.button("🗑️", key=f"del_{row['id']}", use_container_width=True, help="Delete permanently"):
                            delete_reconciliation(row['id'])
                            st.rerun()
    else:
        st.info("No saved history available.")

    # Audit log viewer
    if st.session_state.current_recon_id:
        with st.expander("📋 Audit Log", expanded=False):
            audit_df = get_audit_log(st.session_state.current_recon_id)
            if not audit_df.empty:
                for _, arow in audit_df.iterrows():
                    st.caption(f"**{arow['action_type']}** — {str(arow['timestamp'])[:16]}")
            else:
                st.caption("No actions logged yet.")

    # ── Overdue Follow-up Reminder ────────────────────────────────────────────
    st.markdown("<hr style='border-color:rgba(255,255,255,0.08);margin:8px 0;'>", unsafe_allow_html=True)
    try:
        _overdue_df = get_overdue_followups(days=7)
        if not _overdue_df.empty:
            st.markdown(f"""
            <div class="overdue-alert">
                <div class="overdue-text">⚠️ {len(_overdue_df)} follow-up(s) overdue</div>
                <div class="overdue-sub">7+ days without response</div>
            </div>""", unsafe_allow_html=True)
            with st.expander("View Overdue Follow-ups", expanded=False):
                for _, od in _overdue_df.iterrows():
                    days_ago = (pd.Timestamp.today() - pd.Timestamp(od['notice_sent_date'])).days
                    st.markdown(
                        f"**{od['vendor_name']}**  \n"
                        f"*{od['company_name']} | {od['period']}*  \n"
                        f"📅 Sent {days_ago} days ago · Status: `{od['status']}`"
                    )
                    st.markdown("---")
        else:
            st.caption("✅ No overdue follow-ups.")
    except Exception as _fu_err:
        st.caption(f"⚠️ Could not load follow-up data: {_fu_err}")



# ==========================================
# HEADER
# ==========================================
st.markdown("""
<div class="app-header-banner">
    <div style="display:flex;align-items:center;gap:14px;">
        <div style="width:46px;height:46px;background:rgba(255,255,255,0.12);border-radius:12px;
                    display:flex;align-items:center;justify-content:center;font-size:24px;
                    border:1px solid rgba(255,255,255,0.2);">🛡️</div>
        <div>
            <span class="header-title-text">GST Reconciliation Tool</span>
            <span class="header-version-pill">Enterprise v9.0</span>
            <div class="header-sub-text">Automated B2B · B2BA · CDNR Matching &amp; Compliance Reporting</div>
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

# ── License status banner ─────────────────────────────────────────────────────
if "lic_banner" in st.session_state:
    _mode, _msg = st.session_state["lic_banner"]
    if _mode == "trial":
        st.warning(f"⏳ **Trial Mode** — {_msg}  |  Enter an activation key to unlock full access.")
        with st.expander("🔑 Enter Activation Key", expanded=False):
            _k = st.text_input("Activation Key", placeholder="XXXX-XXXX-XXXX-XXXX", key="inline_key")
            if st.button("Activate Now", key="inline_activate"):
                if _k:
                    from modules.license_manager import activate_key as _activate
                    _r = _activate(_k)
                    if _r["success"]:
                        st.success(_r["message"])
                        st.session_state.pop("lic_banner", None)
                        st.rerun()
                    else:
                        st.error(_r["message"])
    elif _mode == "active":
        st.success(f"✅ **Licensed** — {_msg}")

# ── Seamless Workspace Navigation Bar ───────────────────────────────────────
# Shows only when user is INSIDE a workflow (not on the main dashboard)
if not st.session_state.get('show_dashboard', True):
    _wf_client  = st.session_state.get('meta_name', '')
    _wf_stage   = st.session_state.get('app_stage', 'setup')
    _wf_period  = st.session_state.get('meta_period', '')
    _wf_fy      = st.session_state.get('meta_fy', '')
    _stage_pill = {'setup': '⚙️ Setup', 'processing': '⏳ Processing', 'results': '📊 Results'}.get(_wf_stage, _wf_stage)
    _client_txt = f"  ·  <b>{_wf_client}</b>  ·  {_wf_period} {_wf_fy}" if _wf_client else ""

    nav_col1, nav_col2 = st.columns([5, 1])
    with nav_col1:
        st.markdown(f"""
        <div style="background:#fff;border:1px solid #E2E8F0;border-radius:10px;
                    padding:9px 16px;display:flex;align-items:center;gap:10px;
                    box-shadow:0 1px 4px rgba(0,0,0,.05);margin-bottom:8px;">
            <span style="font-size:12px;color:#94A3B8;font-weight:700;letter-spacing:.04em;">
                🏠 DASHBOARD
            </span>
            <span style="color:#CBD5E1;font-size:14px;">›</span>
            <span style="font-size:12px;color:#2563EB;font-weight:700;">MODULE 02 · GSTR-2B vs Purchase Register</span>
            <span style="font-size:11px;background:#FFFBEB;color:#D97706;border:1px solid rgba(217,119,6,.2);
                        padding:2px 8px;border-radius:12px;font-weight:700;margin-left:4px;">{_stage_pill}</span>
            <span style="font-size:12px;color:#475569;margin-left:2px;">{_client_txt}</span>
        </div>
        """, unsafe_allow_html=True)
    with nav_col2:
        if st.button("🏠 Dashboard", key="topbar_back_dashboard", use_container_width=True):
            st.session_state['show_dashboard'] = True
            st.rerun()

# ==========================================
# MAIN DASHBOARD — landing page gate
# ==========================================
if st.session_state.get('show_dashboard', True):
    _clicked = render_dashboard()
    if _clicked == "MODULE 02":
        # Show a smooth loading overlay before rerun
        st.markdown("""
        <style>
        .gst-loading-overlay {
            position: fixed; top: 0; left: 0; width: 100vw; height: 100vh;
            background: linear-gradient(135deg, #1a2340 0%, #1e3a8a 100%);
            z-index: 99999; display: flex; flex-direction: column;
            align-items: center; justify-content: center;
            animation: gst-overlay-in 0.18s ease-out;
        }
        @keyframes gst-overlay-in {
            from { opacity: 0; } to { opacity: 1; }
        }
        .gst-loading-logo { font-size: 2.8rem; margin-bottom: 12px; }
        .gst-loading-title {
            color: #fff; font-size: 1.3rem; font-weight: 700;
            letter-spacing: 0.5px; margin-bottom: 6px;
        }
        .gst-loading-sub { color: #93c5fd; font-size: 0.85rem; margin-bottom: 28px; }
        .gst-spinner {
            width: 36px; height: 36px; border: 3px solid rgba(255,255,255,0.2);
            border-top-color: #60a5fa; border-radius: 50%;
            animation: gst-spin 0.7s linear infinite;
        }
        @keyframes gst-spin { to { transform: rotate(360deg); } }
        </style>
        <div class="gst-loading-overlay">
            <div class="gst-loading-logo">🛡️</div>
            <div class="gst-loading-title">Opening Reconciliation Workspace</div>
            <div class="gst-loading-sub">Module 02 · GSTR-2B vs Purchase Register</div>
            <div class="gst-spinner"></div>
        </div>
        """, unsafe_allow_html=True)
        import time as _time; _time.sleep(0.35)
        st.session_state['show_dashboard'] = False
        st.session_state['app_stage'] = 'setup'
        st.rerun()
    st.stop()   # ← everything below only runs when dashboard is hidden

# ==========================================
# GSTR-2B MULTI-FILE MERGER (TOP CORNER)
# ==========================================

import copy as _copy
import openpyxl as _openpyxl

# NIC GSTR-2B: row number where actual data starts (everything before = header template)
_GSTR2B_DATA_START = {
    'B2B': 7, 'B2BA': 8, 'B2B-CDNR': 7, 'B2B-CDNRA': 8,
    'IMPG': 7, 'IMPGSEZ': 7, 'ISD': 7, 'ISDA': 8,
    'ECOMM': 7, 'ECOMMA': 8,
    'B2B(REJECTED)': 7, 'B2BA(REJECTED)': 8,
    'B2B-CDNR(REJECTED)': 7, 'B2B-CDNRA(REJECTED)': 8,
    'ECO(REJECTED)': 7, 'ECOA(REJECTED)': 8, 'ISD(REJECTED)': 7,
}

def _copy_ws_row(src_ws, dst_ws, src_row, dst_row):
    """Copy one row with cell values + styles from src to dst worksheet."""
    for col_idx, src_cell in enumerate(src_ws[src_row], 1):
        dst_cell = dst_ws.cell(row=dst_row, column=col_idx)
        dst_cell.value = src_cell.value
        if src_cell.has_style:
            dst_cell.font      = _copy.copy(src_cell.font)
            dst_cell.fill      = _copy.copy(src_cell.fill)
            dst_cell.border    = _copy.copy(src_cell.border)
            dst_cell.alignment = _copy.copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format

def _copy_merged_cells(src_ws, dst_ws, up_to_row):
    """Copy merged cell regions that lie within header rows."""
    for rng in src_ws.merged_cells.ranges:
        if rng.min_row <= up_to_row:
            try:
                dst_ws.merge_cells(str(rng))
            except Exception:
                pass

def _get_sheet(wb, norm_name):
    """Find a sheet in workbook by normalised (upper) name."""
    for s in wb.sheetnames:
        if s.strip().upper() == norm_name:
            return s
    return None

def _extract_data_rows(wb, norm_name, data_start):
    """Return list-of-tuples for data rows in the given sheet."""
    sheet = _get_sheet(wb, norm_name)
    if sheet is None:
        return []
    ws = wb[sheet]
    rows = []
    for r in range(data_start, ws.max_row + 1):
        row = tuple(ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1))
        if any(v is not None for v in row):
            rows.append(row)
    return rows

def merge_gstr2b_files(uploaded_files):
    """
    Properly merges multiple NIC-format GSTR-2B Excel files.
    - Preserves rows 1–(data_start-1) from the first file as template (title, sub-headers, merged cells)
    - Stacks data rows from all files, deduplicating by GSTIN + Invoice No + Date
    Returns (bytes, error_message).
    """
    workbooks = []
    for f in uploaded_files:
        try:
            f.seek(0)
            workbooks.append(_openpyxl.load_workbook(f, data_only=True))
        except Exception as e:
            st.warning(f"⚠️ Could not open `{f.name}`: {e}")

    if not workbooks:
        return None, "No readable GSTR-2B files."

    template_wb = workbooks[0]
    out_wb = _openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    # Collect sheet order from first file, keep same sequence
    for sheet_name in template_wb.sheetnames:
        norm = sheet_name.strip().upper()
        data_start = _GSTR2B_DATA_START.get(norm)
        src_ws = template_wb[sheet_name]
        dst_ws = out_wb.create_sheet(title=sheet_name)

        # Copy column widths
        for col_ltr, dim in src_ws.column_dimensions.items():
            dst_ws.column_dimensions[col_ltr].width = dim.width or 14
        for row_idx, dim in src_ws.row_dimensions.items():
            if dim.height:
                dst_ws.row_dimensions[row_idx].height = dim.height

        if data_start is None:
            # Non-data sheet (Read Me, ITC Available…) — copy as-is from template
            for r in range(1, src_ws.max_row + 1):
                _copy_ws_row(src_ws, dst_ws, r, r)
            _copy_merged_cells(src_ws, dst_ws, src_ws.max_row)
        else:
            # 1 — Copy header rows exactly from template
            for r in range(1, data_start):
                _copy_ws_row(src_ws, dst_ws, r, r)
            _copy_merged_cells(src_ws, dst_ws, data_start - 1)

            # 2 — Collect + deduplicate data from all workbooks
            seen_keys = set()
            all_data  = []
            for wb in workbooks:
                for row in _extract_data_rows(wb, norm, data_start):
                    # Dedup key: col0=GSTIN, col2=Invoice/Note No, col4=Date
                    try:
                        key = (str(row[0]).strip().upper(),
                               str(row[2]).strip().upper(),
                               str(row[4]).strip())
                    except Exception:
                        key = str(row[:5])
                    if key not in seen_keys:
                        seen_keys.add(key)
                        all_data.append(row)

            # 3 — Write merged data
            for offset, data_row in enumerate(all_data):
                write_row = data_start + offset
                for col_idx, val in enumerate(data_row, 1):
                    dst_ws.cell(row=write_row, column=col_idx).value = val

    buf = io.BytesIO()
    out_wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), None


# ── MERGER UI — top right corner via columns ─────────────────────────────────
_merger_col, _merger_btn_col = st.columns([5, 1])
with _merger_btn_col:
    if st.button("🔀 Merge 2B Files", help="Merge multiple GSTR-2B files into one", use_container_width=True):
        st.session_state['show_merger'] = not st.session_state.get('show_merger', False)

if st.session_state.get('show_merger', False):
    with st.container():
        st.markdown("""
        <div style='background: linear-gradient(135deg, #1a237e, #1565c0); color:white;
                    padding:16px 24px; border-radius:12px; margin-bottom:16px;'>
            <h3 style='margin:0;color:white;'>🔀 GSTR-2B Multi-File Merger</h3>
            <p style='margin:4px 0 0 0; font-size:13px; opacity:0.85;'>
                Combine multiple months' GSTR-2B files into a single consolidated Excel.
                Automatically deduplicates by GSTIN + Invoice No + Taxable Value.
            </p>
        </div>
        """, unsafe_allow_html=True)

        merger_files = st.file_uploader(
            "Upload 2 or more GSTR-2B Excel files",
            type=['xlsx'],
            accept_multiple_files=True,
            key="gstr2b_merger_uploader",
            help="Upload GSTR-2B files downloaded from GST Portal (standard NIC format)"
        )

        if merger_files and len(merger_files) >= 2:
            m1, m2, m3 = st.columns(3)
            m1.metric("Files Selected", len(merger_files))
            total_size = sum(f.size for f in merger_files) / 1024
            m2.metric("Total Size", f"{total_size:.1f} KB")
            m3.metric("Format", "GSTR-2B NIC Excel")

            st.markdown("**Files to merge:**")
            for f in merger_files:
                st.caption(f"📄 {f.name}  ({f.size/1024:.1f} KB)")

            if st.button("▶️ Run Merge & Download", type="primary", use_container_width=True, key="run_merger"):
                with st.spinner("Merging files... deduplicating invoices..."):
                    merged_bytes, err = merge_gstr2b_files(merger_files)
                if err:
                    st.error(f"❌ Merge failed: {err}")
                elif merged_bytes:
                    st.success(f"✅ Merged {len(merger_files)} files successfully! Duplicates removed automatically.")
                    st.download_button(
                        label="📥 Download Merged GSTR-2B",
                        data=merged_bytes,
                        file_name=f"GSTR2B_Merged_{len(merger_files)}files.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True
                    )
        elif merger_files and len(merger_files) == 1:
            st.info("☝️ Please upload at least 2 GSTR-2B files to merge.")

        st.markdown("---")

if st.session_state.app_stage == 'setup':

    # Step wizard header
    st.markdown(f"""
    <div style="background:#fff;border-radius:20px;padding:16px 24px;display:flex;align-items:center;
                gap:0;margin-bottom:16px;box-shadow:0 4px 24px rgba(0,0,0,.07)">
      <div style="display:flex;align-items:center;gap:9px">
        <div style="width:28px;height:28px;border-radius:50%;background:#1B2035;color:#fff;display:grid;place-items:center;font-size:11px;font-weight:800">1</div>
        <span style="font-size:12px;font-weight:700;color:#1B2035">Upload Files</span>
      </div>
      <div style="flex:1;height:2px;background:#EDEAE4;margin:0 12px"></div>
      <div style="display:flex;align-items:center;gap:9px">
        <div style="width:28px;height:28px;border-radius:50%;background:#EDEAE4;color:#A8ABBB;display:grid;place-items:center;font-size:11px;font-weight:800">2</div>
        <span style="font-size:12px;font-weight:600;color:#A8ABBB">Column Mapping</span>
      </div>
      <div style="flex:1;height:2px;background:#EDEAE4;margin:0 12px"></div>
      <div style="display:flex;align-items:center;gap:9px">
        <div style="width:28px;height:28px;border-radius:50%;background:#EDEAE4;color:#A8ABBB;display:grid;place-items:center;font-size:11px;font-weight:800">3</div>
        <span style="font-size:12px;font-weight:600;color:#A8ABBB">Settings &amp; Run</span>
      </div>
    </div>
    """, unsafe_allow_html=True)

    with st.container():
        # ── Software selector (compact) ───────────────────────────────────────
        software_names = list(SOFTWARE_COLUMN_PROFILES.keys())
        selected_software = st.selectbox(
            "📋 Accounting Software (for auto column-mapping)",
            software_names,
            index=software_names.index(st.session_state.get('software_profile', software_names[0]))
                  if st.session_state.get('software_profile') in software_names else 0,
            key="software_profile",
            help="Columns are auto-mapped based on your software's export format."
        )
        software_profile = SOFTWARE_COLUMN_PROFILES[selected_software]

        # ── File Upload — clean, focused ──────────────────────────────────────
        st.markdown(f"""
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;margin:14px 0 6px">
          <div style="background:#FFF8E1;border:2px solid #F2C521;border-radius:14px;padding:16px 18px">
            <div style="font-size:13px;font-weight:800;color:#7A5200;margin-bottom:4px">📚 Purchase Register</div>
            <div style="font-size:11px;color:#C89000">From {selected_software}</div>
          </div>
          <div style="background:#F0F4FF;border:2px solid #93C5FD;border-radius:14px;padding:16px 18px">
            <div style="font-size:13px;font-weight:800;color:#1D4ED8;margin-bottom:4px">🏛️ GSTR-2B Portal Data</div>
            <div style="font-size:11px;color:#3B82F6">Download from GST Portal (NIC format)</div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        col1, col2 = st.columns(2)
        with col1:
            file_books = st.file_uploader("Purchase Register", type=['xlsx','csv'],
                                          key="b_up", label_visibility="collapsed")
        with col2:
            file_gst   = st.file_uploader("GSTR-2B Portal Data", type=['xlsx','csv'],
                                          key="g_up", label_visibility="collapsed")

    if file_books and file_gst:

        # Save bytes for CDNR tab
        st.session_state['file_books_bytes'] = file_books.read(); file_books.seek(0)
        st.session_state['file_gst_bytes']   = file_gst.read();   file_gst.seek(0)

        st.divider()
        final_books_map = {}
        final_gst_map   = {}

        # Load data
        df_b_raw = load_data_preview(file_books)
        df_g_raw = load_data_preview(file_gst)

        # --- DATA CONFIDENCE PANEL ---
        if df_b_raw is not None and df_g_raw is not None:
            st.markdown('<div class="section-hdr">📊 Data Confidence Check</div>', unsafe_allow_html=True)
            st.markdown('<div class="section-sub">Verify numbers match your source files before running reconciliation.</div>', unsafe_allow_html=True)
            b_summary = make_data_summary(df_b_raw, "Purchase Register (Books)")
            g_summary = make_data_summary(df_g_raw, "GSTR-2B Portal")
            conf_c1, conf_c2 = st.columns(2)
            with conf_c1:
                st.markdown(f"""
                <div class="confidence-box">
                  <div class="confidence-title">📚 {b_summary['label']}</div>
                  Rows: <b>{b_summary['rows']}</b> &nbsp;|&nbsp;
                  Unique Invoices: <b>{b_summary['invoices']}</b> &nbsp;|&nbsp;
                  Unique GSTINs: <b>{b_summary['gstins']}</b><br>
                  Taxable Value: <b>&#8377; {b_summary['taxable']:,.2f}</b> &nbsp;|&nbsp;
                  IGST: <b>&#8377; {b_summary['igst']:,.2f}</b> &nbsp;|&nbsp;
                  CGST: <b>&#8377; {b_summary['cgst']:,.2f}</b> &nbsp;|&nbsp;
                  SGST: <b>&#8377; {b_summary['sgst']:,.2f}</b><br>
                  <b>Total GST: &#8377; {b_summary['total_gst']:,.2f}</b>
                </div>""", unsafe_allow_html=True)
            with conf_c2:
                st.markdown(f"""
                <div class="confidence-box-gst">
                  <div class="confidence-title-gst">🏛️ {g_summary['label']}</div>
                  Rows: <b>{g_summary['rows']}</b> &nbsp;|&nbsp;
                  Unique Invoices: <b>{g_summary['invoices']}</b> &nbsp;|&nbsp;
                  Unique GSTINs: <b>{g_summary['gstins']}</b><br>
                  Taxable Value: <b>&#8377; {g_summary['taxable']:,.2f}</b> &nbsp;|&nbsp;
                  IGST: <b>&#8377; {g_summary['igst']:,.2f}</b> &nbsp;|&nbsp;
                  CGST: <b>&#8377; {g_summary['cgst']:,.2f}</b> &nbsp;|&nbsp;
                  SGST: <b>&#8377; {g_summary['sgst']:,.2f}</b><br>
                  <b>Total GST: &#8377; {g_summary['total_gst']:,.2f}</b>
                </div>""", unsafe_allow_html=True)
            st.session_state['data_summary_books'] = b_summary
            st.session_state['data_summary_gst']   = g_summary
            st.divider()

        # B2BA amendments
        file_gst.seek(0)
        df_b2ba, status_msg = smart_read_b2ba(file_gst)
        if df_b2ba is not None and not df_b2ba.empty:
            st.info(f"⚡ Processing B2B Amendments... Found {len(df_b2ba)} entries in B2BA.")
            df_g_raw, deleted_count, added_count = process_amendments(df_g_raw, df_b2ba)
            st.success(f"✅ B2B Amendments Applied: Removed {deleted_count} old invoices, Added {added_count} revised.")
        elif status_msg and "Critical" in str(status_msg):
            st.warning(status_msg)

        file_gst.seek(0)
        try:
            _xls_check = pd.ExcelFile(file_gst)
            _has_cdnr  = any('cdnr' in s.lower() for s in _xls_check.sheet_names)
        except Exception:
            _has_cdnr  = False
        file_gst.seek(0)
        if _has_cdnr:
            st.info("📋 CDNR sheet detected in GSTR-2B. Run **CDNR Reconciliation** from **Tab 2** after B2B recon.")

        # Auto-detect metadata
        det_fy, det_period, det_gstin, det_name = "2025 - 2026", "April", "", ""
        meta_fy, meta_period, meta_gstin, meta_name = extract_meta_from_readme(file_gst)
        if meta_gstin: det_gstin  = meta_gstin
        if meta_name:  det_name   = meta_name
        if meta_fy:    det_fy     = meta_fy
        if meta_period: det_period = meta_period

        # Column Mapper
        with st.expander("🛠️ Column Mapping Configuration", expanded=False):
            st.caption(f"Auto-mapped using **{selected_software}** profile. Verify and adjust if needed.")
            cols_books            = list(df_b_raw.columns) if df_b_raw is not None else []
            cols_gst              = list(df_g_raw.columns) if df_g_raw is not None else []
            cols_books_with_blank = ["<No Column / Blank>"] + cols_books

            def smart_find_with_profile(field, available_cols, profile_aliases, fallback_map):
                profile_candidates = profile_aliases.get(field, [])
                available_lower = [str(c).lower() for c in available_cols]
                for candidate in profile_candidates:
                    if candidate in available_cols:
                        return candidate
                    try:
                        idx = available_lower.index(candidate.lower())
                        return available_cols[idx]
                    except ValueError:
                        pass
                    for i, col_l in enumerate(available_lower):
                        if candidate.lower() in col_l or col_l in candidate.lower():
                            return available_cols[i]
                return find_best_match(field, available_cols, fallback_map)

            map_col1, map_col2 = st.columns(2)
            with map_col1:
                st.markdown(f"#### 📚 Purchase Register ({selected_software})")
                for field in REQUIRED_FIELDS:
                    suggested = smart_find_with_profile(field, cols_books_with_blank, software_profile, FIXED_BOOKS_MAPPING)
                    idx = cols_books_with_blank.index(suggested) if suggested in cols_books_with_blank else 0
                    is_profile_match = suggested != "<No Column / Blank>" and suggested in cols_books_with_blank
                    label = f"{field} {'✅' if is_profile_match else '⚠️'}"
                    final_books_map[field] = st.selectbox(label, cols_books_with_blank, index=idx, key=f"b_{field}")
            with map_col2:
                st.markdown("#### 🏛️ GSTR-2B (Portal)")
                for field in REQUIRED_FIELDS:
                    suggested = find_best_match(field, cols_gst, FIXED_GST_MAPPING)
                    idx = cols_gst.index(suggested) if suggested in cols_gst else 0
                    final_gst_map[field] = st.selectbox(f"{field} (GST)", cols_gst, index=idx, key=f"g_{field}")

            mapped_count = sum(1 for v in final_books_map.values() if v != "<No Column / Blank>")
            total_fields = len(REQUIRED_FIELDS)
            confidence_pct = int(mapped_count / total_fields * 100)
            conf_color = "#059669" if confidence_pct >= 80 else "#D97706" if confidence_pct >= 50 else "#C7000A"
            st.markdown(
                f"<div style='margin-top:10px;padding:10px 16px;background:#F8FAFC;"
                f"border-radius:9px;border:1px solid #E2E8F0;'>"
                f"Mapping confidence: <strong style='color:{conf_color}'>"
                f"{confidence_pct}% ({mapped_count}/{total_fields} fields mapped)</strong></div>",
                unsafe_allow_html=True
            )

        st.divider()
        st.markdown('<div class="section-hdr">⚙️ Step 3 — Reconciliation Settings</div>', unsafe_allow_html=True)
        st.markdown('<div class="section-sub">Enter client details and tolerance settings</div>', unsafe_allow_html=True)

        c1, c2, c3, c4 = st.columns(4)
        with c1: gstin_input  = st.text_input("GSTIN", det_gstin)
        with c2: name_input   = st.text_input("Client Name", det_name)
        with c3: fy_input     = st.text_input("Financial Year", det_fy)
        with c4: period_input = st.text_input("Period", det_period)

        # GSTIN Validation
        if gstin_input and not validate_gstin(gstin_input):
            st.warning(f"⚠️ GSTIN format looks invalid: `{gstin_input}`. Expected format: 22AAAAA0000A1Z5")

        t1, t2, t3 = st.columns([1, 1, 2])
        with t1: tolerance_input = st.number_input("Global Tolerance (₹)", min_value=0.0, value=5.0, step=1.0, help="Default allowable difference for matching. Cannot be negative.")
        with t2: smart_mode_input = st.checkbox("Enable Smart Suggestions (Fuzzy Logic)", value=False)

        # Per-vendor tolerance
        with st.expander("⚙️ Per-Vendor Tolerance Overrides (Advanced)", expanded=False):
            st.caption("Override the global tolerance for specific suppliers (GSTIN). Leave empty to use global.")
            vendor_tol_data = st.data_editor(
                pd.DataFrame(list(st.session_state.vendor_tolerances.items()) or [('', 5.0)],
                             columns=['GSTIN', 'Tolerance (₹)']),
                num_rows="dynamic", use_container_width=True, key="vendor_tol_editor"
            )
            if st.button("💾 Save Tolerances"):
                new_tol = {}
                for _, row in vendor_tol_data.iterrows():
                    if row['GSTIN'] and str(row['GSTIN']).strip():
                        new_tol[str(row['GSTIN']).strip().upper()] = float(row['Tolerance (₹)'])
                st.session_state.vendor_tolerances = new_tol
                st.success(f"✅ Saved {len(new_tol)} tolerance overrides.")

        # ── Advanced Period Selector (Optional) ─────────────────────────────
        with st.expander("📅 Advanced — Period Filter (Optional)", expanded=False):
            st.markdown("""
            <div style="background:#EEF4FF;border:1px solid #93C5FD;border-radius:9px;
                        padding:10px 14px;margin-bottom:12px;font-size:12px;color:#1352C9;font-weight:500;">
                💡 <b>Optional:</b> Select specific months to include from each file.
                Leave unselected to use <b>all data</b> (default behaviour).
            </div>""", unsafe_allow_html=True)

            _MONTHS = ["Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","Jan","Feb","Mar"]

            # Auto-detect available months from data
            def _get_months_from_df(df):
                for col in df.columns:
                    if 'date' in col.lower() or 'Date' in col:
                        try:
                            parsed = pd.to_datetime(df[col], dayfirst=True, errors='coerce').dropna()
                            if len(parsed) > 0:
                                return sorted(parsed.dt.strftime('%b').unique().tolist(),
                                              key=lambda m: _MONTHS.index(m) if m in _MONTHS else 99)
                        except (ValueError, TypeError):
                            pass
                return []

            _books_months = _get_months_from_df(df_b_raw) if df_b_raw is not None else []
            _gst_months   = _get_months_from_df(df_g_raw)   if df_g_raw is not None else []

            _pc1, _pc2 = st.columns(2)
            with _pc1:
                st.markdown("**📚 Books Period Filter**")
                st.caption("Months detected in Purchase Register")
                _sel_books = st.multiselect(
                    "Select months (Books)",
                    options=_books_months if _books_months else _MONTHS,
                    default=[],
                    key="period_sel_books",
                    label_visibility="collapsed",
                    placeholder="All months (default)"
                )
            with _pc2:
                st.markdown("**🏛️ GSTR-2B Period Filter**")
                st.caption("Months detected in GSTR-2B")
                _sel_gst = st.multiselect(
                    "Select months (2B)",
                    options=_gst_months if _gst_months else _MONTHS,
                    default=[],
                    key="period_sel_gst",
                    label_visibility="collapsed",
                    placeholder="All months (default)"
                )

            if _sel_books or _sel_gst:
                _info_parts = []
                if _sel_books: _info_parts.append(f"📚 Books: **{', '.join(_sel_books)}**")
                if _sel_gst:   _info_parts.append(f"🏛️ 2B: **{', '.join(_sel_gst)}**")
                st.info(" · ".join(_info_parts) + " — will be filtered before reconciliation.")

        # ── Old ITC Detection Toggle ─────────────────────────────────────────
        st.markdown("---")
        col_itc1, col_itc2 = st.columns([1, 2])
        with col_itc1:
            old_itc_enabled = st.toggle(
                "🗓️ Enable Old ITC Detection",
                value=st.session_state.get('old_itc_enabled', False),
                key='old_itc_toggle',
                help="If ON: any invoice in GSTR-2B whose date falls BEFORE the reconciliation period start date will be tagged as 'Old ITC (Previous Year)' instead of 'Not in Purchase Books'."
            )
            st.session_state['old_itc_enabled'] = old_itc_enabled
        with col_itc2:
            if old_itc_enabled:
                st.info(
                    "🟣 **Old ITC Detection is ON** — Invoices uploaded by supplier in GSTR-1 this year "
                    "whose date falls before the period start (01-Apr of the FY) will be tagged as "
                    "**Old ITC (Previous Year)** and separated from regular discrepancies.",
                    icon=None
                )
            else:
                st.caption("Old ITC Detection is OFF. All 'Not in Books' entries will be treated normally.")
        st.markdown("---")

        if st.button("🚀 Run Reconciliation Engine", type="primary", use_container_width=True):

            # ── Required column mapping validation ─────────────────────────
            _critical_fields = ["GSTIN of Supplier", "Invoice Number", "Taxable Value"]
            _missing_critical = [f for f in _critical_fields if final_books_map.get(f) == "<No Column / Blank>"]
            if _missing_critical:
                st.error(
                    f"❌ Cannot run reconciliation — the following critical columns are not mapped in your "
                    f"Purchase Register: **{', '.join(_missing_critical)}**. "
                    f"Please open the Column Mapping section above and assign these fields before running."
                )
                st.stop()

            if not gstin_input or not gstin_input.strip():
                st.error("❌ Please enter the client's GSTIN in the Settings section above.")
                st.stop()

            if not name_input or not name_input.strip():
                st.error("❌ Please enter the client name in the Settings section above.")
                st.stop()

            # Validate GSTINs in data
            if df_b_raw is not None:
                gstin_col_b = next((c for c in df_b_raw.columns if 'gstin' in c.lower()), None)
                if gstin_col_b:
                    invalid_gstins = df_b_raw[~df_b_raw[gstin_col_b].astype(str).apply(validate_gstin)][gstin_col_b].unique()
                    if len(invalid_gstins) > 0:
                        st.warning(f"⚠️ {len(invalid_gstins)} invalid GSTIN(s) found in Books data. They will be processed but may not match correctly.")

            books_rename_map = {v: k for k, v in final_books_map.items() if v != "<No Column / Blank>"}
            gst_rename_map   = {v: k for k, v in final_gst_map.items()}

            df_b_clean = df_b_raw.rename(columns=books_rename_map)
            df_g_clean = df_g_raw.rename(columns=gst_rename_map)

            df_b_clean = standardize_invoice_numbers(df_b_clean, "Invoice Number")
            df_g_clean = standardize_invoice_numbers(df_g_clean, "Invoice Number")

            # ── Apply period filter if selected ──────────────────────────────
            _sel_books_saved = st.session_state.get('period_sel_books', [])
            _sel_gst_saved   = st.session_state.get('period_sel_gst',   [])
            if _sel_books_saved:
                _date_col_b = next((c for c in df_b_clean.columns if 'date' in c.lower()), None)
                if _date_col_b:
                    _dates_b = pd.to_datetime(df_b_clean[_date_col_b], dayfirst=True, errors='coerce')
                    df_b_clean = df_b_clean[_dates_b.dt.strftime('%b').isin(_sel_books_saved)]
            if _sel_gst_saved:
                _date_col_g = next((c for c in df_g_clean.columns if 'date' in c.lower()), None)
                if _date_col_g:
                    _dates_g = pd.to_datetime(df_g_clean[_date_col_g], dayfirst=True, errors='coerce')
                    df_g_clean = df_g_clean[_dates_g.dt.strftime('%b').isin(_sel_gst_saved)]

            for req_field, mapped_val in final_books_map.items():
                if mapped_val == "<No Column / Blank>":
                    df_b_clean[req_field] = np.nan

            df_b_clean = df_b_clean[[k for k in REQUIRED_FIELDS.keys() if k in df_b_clean.columns]]
            df_g_clean = df_g_clean[[k for k in REQUIRED_FIELDS.keys() if k in df_g_clean.columns]]
            df_b_clean = df_b_clean.loc[:, ~df_b_clean.columns.duplicated()]
            df_g_clean = df_g_clean.loc[:, ~df_g_clean.columns.duplicated()]

            st.session_state['df_b_clean']  = df_b_clean
            st.session_state['df_g_clean']  = df_g_clean
            st.session_state['tolerance']   = tolerance_input
            st.session_state['smart_mode']  = smart_mode_input
            st.session_state['meta_gstin']  = gstin_input
            st.session_state['meta_name']   = name_input
            st.session_state['meta_fy']     = fy_input
            st.session_state['meta_period'] = period_input
            st.session_state['old_itc_enabled'] = old_itc_enabled
            st.session_state.cdnr_result    = None
            st.session_state.cdnr_summary   = None
            st.session_state.current_recon_id = None
            st.session_state['hub_names_skipped'] = False
            st.session_state['hub_names_done']    = False
            st.session_state['combined_report_bytes'] = None
            st.session_state['notices_sent_count'] = 0
            st.session_state.app_stage = 'processing'
            st.rerun()

# ==========================================
# STAGE 2 — PROCESSING
# ==========================================
elif st.session_state.app_stage == 'processing':
    st.markdown("<br><br>", unsafe_allow_html=True)
    st.markdown("<h3 style='text-align:center;color:#444;'>🤖 The Reconciliation Engine is processing your data...</h3>", unsafe_allow_html=True)
    show_processing_animation()

    df_b  = st.session_state['df_b_clean']
    df_g  = st.session_state['df_g_clean']
    tol   = st.session_state['tolerance']
    smart = st.session_state['smart_mode']

    time.sleep(0.5)
    result, df_b_rem, df_g_rem = run_reconciliation(df_b, df_g, tol, st.session_state.manual_matches, smart)
    result['Final_Taxable'] = result['Taxable Value_BOOKS'].fillna(result['Taxable Value_GST']).fillna(0)

    # ── Old ITC Detection (post-processing, non-destructive) ─────────────────
    if st.session_state.get('old_itc_enabled', False):
        fy_str = st.session_state.get('meta_fy', '2025 - 2026')
        try:
            fy_start_year = int(str(fy_str).split('-')[0].strip().split('/')[-1].strip())
        except (ValueError, IndexError):
            fy_start_year = 2025
        period_start = pd.Timestamp(f"{fy_start_year}-04-01")

        # Only tag "Invoices Not in Purchase Books" rows where invoice date < period start
        not_in_books_mask = (
            (result['Recon_Status'] == 'Invoices Not in Purchase Books') &
            (pd.to_datetime(result['Invoice Date_GST'], errors='coerce') < period_start)
        )
        old_itc_count = not_in_books_mask.sum()
        if old_itc_count > 0:
            result.loc[not_in_books_mask, 'Recon_Status'] = 'Old ITC (Previous Year)'
            import streamlit as _st
            _st.info(f"🟣 Old ITC Detection: **{old_itc_count} invoice(s)** re-tagged as 'Old ITC (Previous Year)' "
                     f"(portal date before {period_start.strftime('%d-%m-%Y')}).")

    meta = {
        'gstin':  st.session_state['meta_gstin'],
        'name':   st.session_state['meta_name'],
        'fy':     st.session_state['meta_fy'],
        'period': st.session_state['meta_period']
    }
    recon_id = save_reconciliation(meta, result)
    st.session_state.current_recon_id   = recon_id
    st.session_state.current_client_path = get_client_path(meta['name'], meta['gstin'], meta['fy'], meta['period'])
    st.session_state['last_result'] = result
    log_action(recon_id, 'new_recon', {'invoices': len(result), 'tolerance': tol})
    st.session_state.app_stage = 'results'
    st.rerun()

# ==========================================
# STAGE 3 — RESULTS
# ==========================================
elif st.session_state.app_stage == 'results':

    # ── Result header banner ─────────────────────────────────────────────────
    st.markdown(f"""
    <div class="recon-success-bar" style="display:flex;align-items:center;justify-content:space-between">
        <div style="display:flex;align-items:center;gap:9px">
          <div class="recon-success-dot"></div>
          <div class="recon-success-text">
            ✅ B2B Reconciliation Complete — <b>{st.session_state['meta_name']}</b>
            &nbsp;·&nbsp; {st.session_state['meta_period']} {st.session_state['meta_fy']}
            &nbsp;·&nbsp; <span style="font-family:monospace;font-size:10px">{st.session_state['meta_gstin']}</span>
          </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    _sb_c1, _sb_c2, _sb_c3 = st.columns([5, 1, 1])
    with _sb_c2:
        if st.button("📁 Open Folder", use_container_width=True):
            if st.session_state.current_client_path:
                open_folder(st.session_state.current_client_path)
    with _sb_c3:
        if st.button("🔄 New Recon", type="primary", use_container_width=True):
            st.session_state.app_stage = 'setup'
            st.rerun()

    result = st.session_state['last_result']
    df_b   = st.session_state['df_b_clean']
    df_g   = st.session_state['df_g_clean']
    gstin  = st.session_state['meta_gstin']
    name   = st.session_state['meta_name']
    fy     = st.session_state['meta_fy']
    period = st.session_state['meta_period']

    # Safe display copy
    result_display = result.copy()
    if 'Invoice Date_BOOKS' in result_display.columns:
        result_display['Invoice Date_BOOKS'] = pd.to_datetime(
            result_display['Invoice Date_BOOKS'], dayfirst=True, errors='coerce'
        ).dt.strftime('%d/%m/%Y').fillna(result_display['Invoice Date_BOOKS'])
    if 'Invoice Date_GST' in result_display.columns:
        result_display['Invoice Date_GST'] = pd.to_datetime(
            result_display['Invoice Date_GST'], dayfirst=True, errors='coerce'
        ).dt.strftime('%d/%m/%Y').fillna(result_display['Invoice Date_GST'])
    # Ensure Match_Confidence exists (backward compat with history-loaded results)
    if 'Match_Confidence' not in result_display.columns:
        def _backfill_confidence(row):
            st = str(row.get('Recon_Status',''))
            ml = str(row.get('Match_Logic',''))
            if 'Exact' in ml:        return 100.0
            if 'Date Mismatch' in ml: return 92.0
            if 'Invoice Mismatch' in ml: return 85.0
            if 'Value Mismatch' in ml:   return 78.0
            if 'Group Match' in st:   return 60.0
            if 'Suggestion' in st:    return 65.0
            if 'Matched' in st:       return 100.0
            return 0.0
        result_display['Match_Confidence'] = result_display.apply(_backfill_confidence, axis=1)

    # ─────────────────────────────────────────────────────
    # WORKFLOW PROGRESS BANNER
    # ─────────────────────────────────────────────────────
    _cdnr_done    = st.session_state.get('cdnr_result') is not None
    _combined_rdy = st.session_state.get('combined_report_bytes') is not None
    _notices_sent = st.session_state.get('notices_sent_count', 0)

    def _wf_step(num, label, done, active=False):
        if done:
            _circle = f'<div style="width:28px;height:28px;border-radius:50%;background:#059669;color:#fff;display:flex;align-items:center;justify-content:center;font-size:12px;font-weight:800;flex-shrink:0">&#10003;</div>'
            _txt    = f'<span style="font-size:12px;font-weight:700;color:#059669">{label}</span>'
        elif active:
            _circle = f'<div style="width:28px;height:28px;border-radius:50%;background:#1352C9;color:#fff;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:800;flex-shrink:0;animation:gst-pulse 1.6s ease-in-out infinite">{num}</div>'
            _txt    = f'<span style="font-size:13px;font-weight:800;color:#1352C9">{label}</span>'
        else:
            _circle = f'<div style="width:28px;height:28px;border-radius:50%;background:#E2E8F0;color:#94A3B8;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:700;flex-shrink:0">{num}</div>'
            _txt    = f'<span style="font-size:12px;font-weight:600;color:#94A3B8">{label}</span>'
        return f'<div style="display:flex;align-items:center;gap:8px">{_circle}{_txt}</div>'

    def _wf_line(done):
        bg = '#059669' if done else '#E2E8F0'
        return f'<div style="flex:1;height:2px;background:{bg};margin:0 6px;min-width:20px"></div>'

    _b2b_done = True  # already in results stage
    _dl_done  = _combined_rdy
    _ntc_done = _notices_sent > 0

    if not _cdnr_done:   _active = 2
    elif not _dl_done:   _active = 3
    elif not _ntc_done:  _active = 4
    else:                _active = 5

    st.markdown(f"""
    <div style="background:#fff;border:1px solid #E2E8F0;border-radius:12px;padding:14px 20px;
                margin-bottom:16px;box-shadow:0 1px 4px rgba(0,0,0,.05)">
      <div style="font-size:10px;font-weight:800;color:#94A3B8;letter-spacing:.08em;
                  text-transform:uppercase;margin-bottom:10px">WORKFLOW PROGRESS</div>
      <div style="display:flex;align-items:center">
        {_wf_step(1,'B2B Recon', _b2b_done, _active==1)}
        {_wf_line(_b2b_done)}
        {_wf_step(2,'CDN Recon', _cdnr_done, _active==2)}
        {_wf_line(_cdnr_done)}
        {_wf_step(3,'Downloads', _dl_done, _active==3)}
        {_wf_line(_dl_done)}
        {_wf_step(4,'Send Notices', _ntc_done, _active==4)}
        {_wf_line(_ntc_done)}
        {_wf_step(5,'Follow-up', False, _active==5)}
      </div>
    </div>
    """, unsafe_allow_html=True)

    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8, tab9 = st.tabs([
        "📊 Dashboard & Scorecard",
        "📋 CDNR Matching",
        "📥 Downloads Hub",
        "📝 Detailed Data",
        "🏢 Supplier Wise",
        "🔗 Manual Matcher",
        "💬 Send Notice",
        "📌 Follow-up Tracker",
        "💾 Backup & Restore",
    ])

    # ─────────────────────────────────────────────────────
    # TAB 1 — DASHBOARD
    # ─────────────────────────────────────────────────────
    with tab1:
        total_books_val  = df_b['Taxable Value'].sum() if 'Taxable Value' in df_b.columns else 0
        total_gst_val    = df_g['Taxable Value'].sum() if 'Taxable Value' in df_g.columns else 0
        diff_val         = total_books_val - total_gst_val

        # Compute GST component totals from result
        def _col_sum(df, col):
            return float(pd.to_numeric(df[col], errors='coerce').fillna(0).sum()) if col in df.columns else 0.0
        _books_igst = _col_sum(result, 'IGST_BOOKS'); _books_cgst = _col_sum(result, 'CGST_BOOKS'); _books_sgst = _col_sum(result, 'SGST_BOOKS')
        _gst_igst   = _col_sum(result, 'IGST_GST');   _gst_cgst   = _col_sum(result, 'CGST_GST');   _gst_sgst   = _col_sum(result, 'SGST_GST')
        _books_total_gst = _books_igst + _books_cgst + _books_sgst
        _gst_total_gst   = _gst_igst   + _gst_cgst   + _gst_sgst
        _diff_gst        = _books_total_gst - _gst_total_gst

        # ── ITC Net Summary Banner ─────────────────────────────────────────────
        _not_in_2b_df  = result[result['Recon_Status'] == 'Invoices Not in GSTR-2B']
        _itc_blocked   = float(_not_in_2b_df['Final_Taxable'].sum()) if 'Final_Taxable' in _not_in_2b_df.columns else 0.0
        _cdnr_itc      = float(st.session_state.cdnr_summary.get('net_itc_impact', 0)) if st.session_state.cdnr_summary else 0.0
        _net_eligible  = total_books_val - _itc_blocked + _cdnr_itc

        def _fmt(v):
            return "Rs. {:,.0f}".format(v)

        _banner_html = (
            '<div class="itc-net-banner">'
            '<div class="itc-item">'
            '<div class="itc-item-label">Books Taxable (B2B)</div>'
            '<div class="itc-item-val">' + _fmt(total_books_val) + '</div>'
            '</div>'
            '<div class="itc-sep"></div>'
            '<div class="itc-item">'
            '<div class="itc-item-label">GSTR-2B Taxable</div>'
            '<div class="itc-item-val itc-item-val-blue">' + _fmt(total_gst_val) + '</div>'
            '</div>'
            '<div class="itc-sep"></div>'
            '<div class="itc-item">'
            '<div class="itc-item-label">ITC Blocked (Not in 2B)</div>'
            '<div class="itc-item-val itc-item-val-red">' + _fmt(_itc_blocked) + '</div>'
            '</div>'
        )
        if st.session_state.cdnr_result is not None and st.session_state.cdnr_summary:
            _banner_html += (
                '<div class="itc-sep"></div>'
                '<div class="itc-item">'
                '<div class="itc-item-label">CDNR Adjustment</div>'
                '<div class="itc-item-val">' + _fmt(_cdnr_itc) + '</div>'
                '</div>'
            )
        _banner_html += (
            '<div class="itc-sep"></div>'
            '<div class="itc-item">'
            '<div class="itc-item-label">Net Eligible ITC</div>'
            '<div class="itc-item-val itc-item-val-grn">' + _fmt(_net_eligible) + '</div>'
            '</div>'
            '</div>'
        )
        st.markdown(_banner_html, unsafe_allow_html=True)

        # ── GST component retained for downstream calcs ───────────────────────
        _diff_igst = _books_igst - _gst_igst

        # ── 4 KPI tiles ───────────────────────────────────────────────────────
        _total_inv       = len(result)
        _matched         = int((result['Recon_Status'].str.contains('Matched', na=False)).sum())
        _not_in_2b_count = int((result['Recon_Status'] == 'Invoices Not in GSTR-2B').sum())
        _mismatch        = int((result['Recon_Status'].str.contains('Mismatch', na=False)).sum())

        def _kpi(label, val, bg='#FFFFFF', lc='#A8ABBB', vc='#1B2035'):
            return f"""<div style="background:{bg};border-radius:14px;padding:15px 16px;box-shadow:0 4px 24px rgba(0,0,0,.07)">
              <div style="font-size:9px;font-weight:800;text-transform:uppercase;letter-spacing:.07em;color:{lc};margin-bottom:6px">{label}</div>
              <div style="font-size:22px;font-weight:800;letter-spacing:-.025em;color:{vc}">{val}</div>
            </div>"""

        st.markdown(f"""
        <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:16px">
          {_kpi('Total Invoices', f'{_total_inv:,}')}
          {_kpi('✅ Matched', f'{_matched:,}', '#FFF8E1', '#C89000', '#7A5200')}
          {_kpi('❌ Not in GSTR-2B', f'{_not_in_2b_count:,}', '#FEF0ED', '#A82C18', '#D63920')}
          {_kpi('⚠️ Value Mismatch', f'{_mismatch:,}')}
        </div>
        """, unsafe_allow_html=True)

        # ── Donut + Risk Radar 2-col ──────────────────────────────────────────
        summary_df = result['Recon_Status'].value_counts().reset_index()
        summary_df.columns = ['Status', 'Count']
        _match_pct = int(_matched / max(_total_inv, 1) * 100)
        _not2b_pct = int(_not_in_2b_count / max(_total_inv, 1) * 100)
        _mis_pct   = int(_mismatch / max(_total_inv, 1) * 100)
        _circ = 264
        _matched_dash = int(_match_pct / 100 * _circ)
        _offset_matched = _circ - _matched_dash

        _col_donut, _col_risk = st.columns([1, 1], gap="medium")

        with _col_donut:
            def _bar_row(color, label, count, pct):
                return (
                    f'<div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:2px">'
                    f'<span style="font-size:11px;color:#475569;display:flex;align-items:center;gap:5px">'
                    f'<span style="width:8px;height:8px;background:{color};border-radius:50%;display:inline-block"></span>'
                    f'<b>{label}</b></span><b style="font-size:12px;color:#0D1B40">{count:,}</b></div>'
                    f'<div style="height:6px;background:#F1F5F9;border-radius:50px;margin:3px 0 10px;overflow:hidden">'
                    f'<div style="height:100%;border-radius:50px;width:{pct}%;background:{color}"></div></div>'
                )
            st.markdown(f"""
            <div style="background:#fff;border-radius:20px;padding:20px;box-shadow:0 4px 24px rgba(0,0,0,.07)">
              <div style="font-size:13px;font-weight:700;color:#1B2035;margin-bottom:14px">Match Rate by Status</div>
              <div style="display:flex;align-items:center;gap:20px">
                <div style="position:relative;flex-shrink:0">
                  <svg width="110" height="110" viewBox="0 0 110 110">
                    <circle cx="55" cy="55" r="42" fill="none" stroke="#EDEAE4" stroke-width="12"/>
                    <circle cx="55" cy="55" r="42" fill="none" stroke="#F2C521" stroke-width="12"
                      stroke-dasharray="{_circ}" stroke-dashoffset="{_offset_matched}"
                      stroke-linecap="round" transform="rotate(-90 55 55)"/>
                  </svg>
                  <div style="position:absolute;inset:0;display:flex;flex-direction:column;align-items:center;justify-content:center">
                    <div style="font-size:22px;font-weight:800;color:#1B2035">{_match_pct}%</div>
                    <div style="font-size:9px;color:#A8ABBB;font-weight:700">matched</div>
                  </div>
                </div>
                <div style="flex:1">
                  {_bar_row('#F2C521','Matched',_matched,_match_pct)}
                  {_bar_row('#D63920','Not in 2B',_not_in_2b_count,_not2b_pct)}
                  {_bar_row('#5B6EE8','Mismatch',_mismatch,_mis_pct)}
                </div>
              </div>
            </div>
            """, unsafe_allow_html=True)

        with _col_risk:
            not_in_2b = result[result['Recon_Status'] == 'Invoices Not in GSTR-2B']
            if not not_in_2b.empty:
                risk_df = not_in_2b.groupby('Name of Party').agg(
                    Missing_Count=('GSTIN','count'), Total_Value=('Final_Taxable','sum')
                ).reset_index().sort_values('Total_Value', ascending=False).head(5)
                _max_risk = float(risk_df['Total_Value'].max())
                _risk_rows = ""
                for _ri, (_, _r) in enumerate(risk_df.iterrows(), 1):
                    _rk_pct = int(_r['Total_Value'] / max(_max_risk,1) * 100)
                    _bar_c = '#D63920' if _ri <= 2 else '#F2C521'
                    _nm = str(_r['Name of Party'])[:24]
                    _cnt = int(_r['Missing_Count'])
                    _val = f"&#8377;{_r['Total_Value']:,.0f}"
                    _risk_rows += (
                        f'<div style="display:flex;align-items:center;gap:10px;padding:9px 0;border-bottom:1px solid #F5F2ED">'
                        f'<div style="width:22px;height:22px;border-radius:6px;background:#FEF0ED;color:#D63920;font-size:10px;font-weight:800;display:grid;place-items:center;flex-shrink:0">{_ri}</div>'
                        f'<div style="flex:1"><div style="font-size:12px;font-weight:700;color:#0D1B40">{_nm}</div>'
                        f'<div style="font-size:10px;color:#94A3B8;margin-top:1px">{_cnt} invoices missing from GSTR-2B</div></div>'
                        f'<div style="width:60px"><div style="height:5px;background:#F1F5F9;border-radius:50px;overflow:hidden">'
                        f'<div style="height:100%;border-radius:50px;width:{_rk_pct}%;background:{_bar_c}"></div></div></div>'
                        f'<div style="font-size:12px;font-weight:800;color:#D63920;white-space:nowrap">{_val}</div></div>'
                    )
                st.markdown(f"""
                <div style="background:#fff;border-radius:20px;padding:20px;box-shadow:0 4px 24px rgba(0,0,0,.07)">
                  <div style="font-size:13px;font-weight:700;color:#1B2035;margin-bottom:14px">🚨 Risk Radar — Top Vendors</div>
                  {_risk_rows}
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown(f"""
                <div style="background:#EDFAF3;border-radius:20px;padding:24px;text-align:center;box-shadow:0 4px 24px rgba(0,0,0,.07)">
                  <div style="font-size:28px;margin-bottom:8px">✅</div>
                  <div style="font-size:13px;font-weight:800;color:#18895A">All Clear!</div>
                  <div style="font-size:11px;color:#A8ABBB;margin-top:4px">No vendors with missing B2B invoices</div>
                </div>
                """, unsafe_allow_html=True)

        # ── Data Confidence panel ─────────────────────────────────────────────
        b_sum = st.session_state.get('data_summary_books')
        g_sum = st.session_state.get('data_summary_gst')
        if b_sum or g_sum:
            with st.expander("📋 Data Loaded Summary — Confidence Check", expanded=False):
                dc1, dc2 = st.columns(2)
                if b_sum:
                    dc1.markdown(f"""
                    **📚 Purchase Register (Books)**  
                    Rows: `{b_sum['rows']}` | Invoices: `{b_sum['invoices']}` | GSTINs: `{b_sum['gstins']}`  
                    Taxable: **Rs. {b_sum['taxable']:,.2f}** | IGST: **Rs. {b_sum['igst']:,.2f}**  
                    CGST: **Rs. {b_sum['cgst']:,.2f}** | SGST: **Rs. {b_sum['sgst']:,.2f}**  
                    **Total GST: Rs. {b_sum['total_gst']:,.2f}**
                    """)
                if g_sum:
                    dc2.markdown(f"""
                    **🏛 GSTR-2B Portal**  
                    Rows: `{g_sum['rows']}` | Invoices: `{g_sum['invoices']}` | GSTINs: `{g_sum['gstins']}`  
                    Taxable: **Rs. {g_sum['taxable']:,.2f}** | IGST: **Rs. {g_sum['igst']:,.2f}**  
                    CGST: **Rs. {g_sum['cgst']:,.2f}** | SGST: **Rs. {g_sum['sgst']:,.2f}**  
                    **Total GST: Rs. {g_sum['total_gst']:,.2f}**
                    """)

        # ── Next-step call-to-action ───────────────────────────────────────
        _cdnr_done_t1 = st.session_state.get('cdnr_result') is not None
        if _cdnr_done_t1:
            st.markdown("""
            <div class="next-step-hint">
              <div style="font-size:11px;font-weight:800;color:#1352C9;letter-spacing:.07em;text-transform:uppercase;margin-bottom:5px">
                ✅ Both Recons Done — Next Step
              </div>
              <div style="font-size:14px;font-weight:800;color:#0D1B40">
                <span class="next-step-arrow">→</span> Go to <b>📥 Downloads Hub (Tab 3)</b> to get all your reports in one place
              </div>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown("""
            <div class="next-step-hint">
              <div style="font-size:11px;font-weight:800;color:#1352C9;letter-spacing:.07em;text-transform:uppercase;margin-bottom:5px">
                ✅ B2B Recon Done — Next Step
              </div>
              <div style="font-size:14px;font-weight:800;color:#0D1B40">
                <span class="next-step-arrow">→</span> Click <b>📋 CDNR Matching (Tab 2)</b> above to run CDN Recon
              </div>
            </div>
            """, unsafe_allow_html=True)

    # ─────────────────────────────────────────────────────
    # TAB 3 — DOWNLOADS HUB  (new — all downloads in one place)
    # ─────────────────────────────────────────────────────
    with tab3:
        st.markdown("""
        <div style="background:#0D1B40;border-radius:14px;padding:18px 24px;margin-bottom:18px">
          <div style="font-size:16px;font-weight:800;color:#fff;margin-bottom:3px">📥 Downloads Hub</div>
          <div style="font-size:12px;color:rgba(255,255,255,.5)">All 4 reports generated in sequence — CDNR → B2B → Combined → ITC PDF</div>
        </div>
        """, unsafe_allow_html=True)

        # ── Check for unknown names ─────────────────────────────────────────
        _cdnr_for_names = st.session_state.get('cdnr_result')
        _unknown_in_hub = []
        if _cdnr_for_names is not None and 'Name of Party' in _cdnr_for_names.columns:
            _unk_mask = _cdnr_for_names['Name of Party'].isin(['Unknown', '', 'nan', 'UNKNOWN']) | _cdnr_for_names['Name of Party'].isna()
            _unknown_in_hub = _cdnr_for_names.loc[_unk_mask, 'GSTIN'].dropna().unique().tolist()
            _unknown_in_hub = [g for g in _unknown_in_hub if g and str(g) not in ('', 'nan')]

        # Session flags
        _names_skipped = st.session_state.get('hub_names_skipped', False)
        _names_done    = st.session_state.get('hub_names_done', False)
        _show_downloads = (not _unknown_in_hub) or _names_skipped or _names_done

        # ── STEP 1: NAME UPDATE (only if unknowns exist and not skipped) ────
        if _unknown_in_hub and not _show_downloads:
            st.markdown(f"""
            <div style="background:#FFFBEB;border:2px solid #F59E0B;border-radius:12px;padding:16px 20px;margin-bottom:12px">
              <div style="font-size:14px;font-weight:800;color:#92400E;margin-bottom:4px">
                ✏️ Unknown Vendor Names Detected — {len(_unknown_in_hub)} GSTINs
              </div>
              <div style="font-size:12px;color:#78350F;line-height:1.6">
                Your CDNR data has <b>{len(_unknown_in_hub)} vendors</b> showing as "Unknown".<br>
                Update their names below for accurate reports, or skip to download now.
              </div>
            </div>
            """, unsafe_allow_html=True)

            # ── TRUE SPREADSHEET — st.data_editor ──────────────────────────
            _name_tbl = pd.DataFrame({
                'GST':        _unknown_in_hub,
                'Trade Name': [st.session_state.get(f'cdnr_name_{_g}', '') for _g in _unknown_in_hub]
            })
            _edited_names = st.data_editor(
                _name_tbl,
                column_config={
                    'GST':        st.column_config.TextColumn('GST',        disabled=True,  width='medium'),
                    'Trade Name': st.column_config.TextColumn('Trade Name', disabled=False, width='large'),
                },
                hide_index=True,
                use_container_width=True,
                num_rows='fixed',
                key='hub_name_editor',
            )

            st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
            _btn_col1, _btn_col2 = st.columns([3, 1])
            with _btn_col1:
                if st.button("✅ Update Names & Proceed to Downloads", type="primary", use_container_width=True, key="hub_apply_names"):
                    _updated_hub = 0
                    _new_cdnr_hub = st.session_state.cdnr_result.copy()
                    for _, _row in _edited_names.iterrows():
                        _g = _row['GST']
                        _n = str(_row['Trade Name']).strip()
                        if _n and _n not in ('', 'nan'):
                            _new_cdnr_hub.loc[_new_cdnr_hub['GSTIN'] == _g, 'Name of Party'] = _n
                            st.session_state[f'cdnr_name_{_g}'] = _n
                            _updated_hub += 1
                    if _updated_hub:
                        _last_r = st.session_state.get('last_result')
                        if _last_r is not None and 'GSTIN' in _last_r.columns:
                            for _, _row in _edited_names.iterrows():
                                _g = _row['GST']
                                _n = str(_row['Trade Name']).strip()
                                if _n and _n not in ('', 'nan'):
                                    _mask_l = (_last_r['GSTIN'] == _g) & (_last_r['Name of Party'].isin(['Unknown','','UNKNOWN']))
                                    _last_r.loc[_mask_l, 'Name of Party'] = _n
                            st.session_state['last_result'] = _last_r
                        st.session_state.cdnr_result = _new_cdnr_hub
                        st.session_state['combined_report_bytes'] = None
                        st.session_state['hub_names_done'] = True
                        st.rerun()
                    else:
                        st.warning("No names entered — type trade names in the table above, then click Update.")
            with _btn_col2:
                if st.button("⏭ Skip for Now", use_container_width=True, key="hub_skip_names"):
                    st.session_state['hub_names_skipped'] = True
                    st.rerun()

        else:
            # ── Name status bar ─────────────────────────────────────────────
            if _names_done:
                st.markdown('<div style="background:#F0FDF7;border:1.5px solid #A7F3D0;border-radius:10px;padding:10px 16px;margin-bottom:14px;font-size:12px;font-weight:700;color:#065F46">✅ Vendor names updated — reports will show correct names</div>', unsafe_allow_html=True)
            elif _names_skipped and _unknown_in_hub:
                st.markdown(f'<div style="background:#FFF8E1;border:1.5px solid #FDE68A;border-radius:10px;padding:10px 16px;margin-bottom:14px;font-size:12px;font-weight:700;color:#92400E">⚠️ Skipped name update — {len(_unknown_in_hub)} GSTINs will show as "Unknown" in reports</div>', unsafe_allow_html=True)
            elif not _unknown_in_hub and st.session_state.get('cdnr_result') is not None:
                st.markdown('<div style="background:#F0FDF7;border:1.5px solid #A7F3D0;border-radius:10px;padding:10px 16px;margin-bottom:14px;font-size:12px;font-weight:700;color:#065F46">✅ All vendor names are set — ready to download reports</div>', unsafe_allow_html=True)

            # ── DOWNLOAD SECTION ────────────────────────────────────────────
            _cdnr_rdy  = st.session_state.get('cdnr_result') is not None
            _comb_rdy  = st.session_state.get('combined_report_bytes') is not None
            _both_done = _cdnr_rdy  # B2B is always done if we're here

            # Pre-generate all report bytes
            _cdnr_r     = st.session_state.get('cdnr_result')
            _cdnr_bytes = None
            _b2b_bytes  = None
            _itc_bytes  = None

            if _cdnr_rdy:
                try:
                    _cdnr_bytes = generate_cdnr_excel(_cdnr_r, gstin, name, fy, period, b2b_full_df=result)
                except Exception:
                    _cdnr_bytes = None

            try:
                _b2b_bytes = generate_excel(result, gstin, name, fy, period)
            except Exception:
                _b2b_bytes = None

            try:
                _itc_bytes = create_itc_risk_pdf(result, name, gstin, period, fy).getvalue()
            except Exception:
                _itc_bytes = None

            # Auto-generate Combined
            if _both_done and not _comb_rdy:
                with st.spinner("⚡ Building Combined Report..."):
                    try:
                        _auto_cb = generate_combined_excel(st.session_state['last_result'], st.session_state['cdnr_result'], gstin, name, fy, period)
                        st.session_state['combined_report_bytes'] = _auto_cb
                        _comb_rdy = True
                    except Exception as _ae:
                        st.warning(f"Could not build combined report: {_ae}")

            _comb_bytes = st.session_state.get('combined_report_bytes')

            _XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            _reports = [
                {"n":"1","icon":"📋","title":"CDNR Reconciliation",     "desc":"Credit/Debit Note matching — all statuses",            "fname":f"CDNR_Reconciliation_{period}.xlsx",       "mime":_XLSX,             "bytes":_cdnr_bytes, "key":"dl_cdnr",     "badge":"#1352C9"},
                {"n":"2","icon":"📘","title":"B2B Reconciliation",       "desc":"Full B2B matching with Executive Summary",             "fname":f"B2B_Reconciliation_Report_{period}.xlsx", "mime":_XLSX,             "bytes":_b2b_bytes,  "key":"dl_b2b",      "badge":"#0F6B3C"},
                {"n":"3","icon":"📊","title":"Combined B2B + CDNR",      "desc":"Unified Executive Summary + all sheets in one Excel",  "fname":f"Combined_Reconciliation_{period}.xlsx",   "mime":_XLSX,             "bytes":_comb_bytes, "key":"dl_combined", "badge":"#B45309"},
                {"n":"4","icon":"📑","title":"ITC Risk Summary (PDF)",   "desc":"One-pager ITC at risk — ready to share with your CA", "fname":f"ITC_Risk_Summary_{period}.pdf",           "mime":"application/pdf", "bytes":_itc_bytes,  "key":"dl_itc",      "badge":"#7C3AED"},
            ]

            # Table header
            st.markdown('<div style="display:grid;grid-template-columns:44px 1fr 130px;background:#0D1B40;border-radius:10px 10px 0 0;padding:10px 18px;font-size:10px;font-weight:800;color:rgba(255,255,255,.55);letter-spacing:.07em;text-transform:uppercase"><div>#</div><div>Report</div><div style="text-align:center">Download</div></div>', unsafe_allow_html=True)

            for _rpt in _reports:
                _is_rdy = _rpt["bytes"] is not None
                _st_badge = ('<span style="background:#F0FDF7;color:#0F6B3C;border-radius:20px;padding:2px 9px;font-size:9px;font-weight:800">✅ READY</span>' if _is_rdy else '<span style="background:#FFF1F2;color:#C7000A;border-radius:20px;padding:2px 9px;font-size:9px;font-weight:800">⚠ NOT READY</span>')
                _ci, _cb = st.columns([5, 1])
                with _ci:
                    st.markdown(
                        f'<div style="display:grid;grid-template-columns:44px 1fr;align-items:center;padding:13px 18px;border:1.5px solid #E2E8F0;border-top:none;background:#fff">'
                        f'<div style="width:30px;height:30px;border-radius:8px;background:{_rpt["badge"]}18;display:flex;align-items:center;justify-content:center;font-size:15px">{_rpt["icon"]}</div>'
                        f'<div style="padding-left:12px"><span style="background:{_rpt["badge"]};color:#fff;border-radius:4px;padding:1px 7px;font-size:9px;font-weight:800;font-family:monospace;margin-right:7px">{_rpt["n"]}</span>'
                        f'<b style="font-size:13px;color:#0D1B40">{_rpt["title"]}</b>'
                        f'<div style="font-size:11px;color:#64748B;margin-top:2px">{_rpt["desc"]}</div>'
                        f'<div style="margin-top:5px">{_st_badge}</div></div></div>',
                        unsafe_allow_html=True)
                with _cb:
                    if _is_rdy:
                        st.download_button("📥 Download", data=_rpt["bytes"], file_name=_rpt["fname"], mime=_rpt["mime"],
                                           type="primary", use_container_width=True, key=_rpt["key"],
                                           on_click=save_callback, args=(st.session_state.current_client_path, _rpt["fname"], _rpt["bytes"]))
                    else:
                        st.button("🔒 Not Ready", disabled=True, use_container_width=True, key=f'{_rpt["key"]}_locked')

            st.markdown('<div style="border:1.5px solid #E2E8F0;border-top:none;border-radius:0 0 10px 10px;height:6px;background:#F8FAFC"></div>', unsafe_allow_html=True)
            st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)

            # ── DOWNLOAD ALL AS ZIP ──────────────────────────────────────────
            _ready_rpts = [r for r in _reports if r["bytes"] is not None]
            if _ready_rpts:
                import zipfile as _zf, io as _io_z
                _zbuf = _io_z.BytesIO()
                with _zf.ZipFile(_zbuf, 'w', _zf.ZIP_DEFLATED) as _z:
                    for _r in _ready_rpts:
                        _z.writestr(_r["fname"], _r["bytes"])
                _zbuf.seek(0)
                _zname = f"GST_Reports_{name}_{period}.zip"
                st.markdown(f'<div style="background:#0D1B40;border-radius:12px;padding:14px 20px;margin-bottom:6px"><div style="font-size:14px;font-weight:800;color:#fff">⚡ Download All {len(_ready_rpts)} Reports — One Click</div><div style="font-size:11px;color:rgba(255,255,255,.4);margin-top:2px">All ready reports bundled as a single ZIP file</div></div>', unsafe_allow_html=True)
                st.download_button(f"📦 Download All {len(_ready_rpts)} Reports as ZIP",
                    data=_zbuf.getvalue(), file_name=_zname, mime="application/zip",
                    type="primary", use_container_width=True, key="dl_all_zip")

            # ── Next step ────────────────────────────────────────────────────
            if all(r["bytes"] is not None for r in _reports):
                st.markdown('<div class="next-step-hint" style="margin-top:14px"><div style="font-size:11px;font-weight:800;color:#1352C9;letter-spacing:.07em;text-transform:uppercase;margin-bottom:5px">✅ All Reports Ready — Next Step</div><div style="font-size:14px;font-weight:800;color:#0D1B40"><span class="next-step-arrow">→</span> Click <b>💬 Send Notice (Tab 7)</b> to notify vendors with missing invoices</div></div>', unsafe_allow_html=True)

    # ─────────────────────────────────────────────────────
    # TAB 3 — DETAILED DATA
    # ─────────────────────────────────────────────────────
    with tab4:
        _t3c1, _t3c2 = st.columns([2, 2])
        with _t3c1:
            filters = ["All Data", "Matched", "Mismatch (Value)", "AI Matched",
                       "Suggestions", "🔗 Group Match", "Manually Linked", "Not in 2B", "Not in Books"]
            status_filter = st.selectbox("Filter by Status:", filters, index=0)
        with _t3c2:
            _live_search = st.text_input("🔍 Search vendor / invoice number", placeholder="Type to filter rows...", key="tab3_search")

        df_view = result_display.copy()

        if status_filter == "All Data":
            pass
        elif status_filter == "Matched":
            df_view = result_display[result_display['Recon_Status'].str.contains('Matched', na=False) &
                                     ~result_display['Recon_Status'].str.contains('AI', na=False)]
        elif status_filter == "Mismatch (Value)":
            df_view = result_display[result_display['Recon_Status'].str.contains('Mismatch', na=False)]
        elif status_filter == "AI Matched":
            df_view = result_display[result_display['Recon_Status'].str.contains('AI', na=False)].copy()
            # Show confidence distribution summary
            if not df_view.empty and 'Match_Confidence' in df_view.columns:
                _conf = df_view['Match_Confidence']
                _high   = int((_conf >= 90).sum())
                _medium = int(((_conf >= 75) & (_conf < 90)).sum())
                _low    = int((_conf < 75).sum())
                _avg    = float(_conf.mean())
                st.markdown(f"""
                <div style='display:flex;gap:10px;flex-wrap:wrap;margin-bottom:10px;align-items:center;'>
                  <span style='font-size:12px;font-weight:700;color:#475569;'>🎯 AI Match Confidence:</span>
                  <span style='background:#F0FDF4;color:#166534;padding:3px 12px;border-radius:20px;font-size:12px;font-weight:700;border:1px solid #86EFAC;'>
                    🟢 High (≥90%) &nbsp;{_high}
                  </span>
                  <span style='background:#FFFBEB;color:#92400E;padding:3px 12px;border-radius:20px;font-size:12px;font-weight:700;border:1px solid #FCD34D;'>
                    🟡 Medium (75–89%) &nbsp;{_medium}
                  </span>
                  <span style='background:#FFF1F2;color:#9F1239;padding:3px 12px;border-radius:20px;font-size:12px;font-weight:700;border:1px solid #FDA4AF;'>
                    🔴 Low (<75%) &nbsp;{_low}
                  </span>
                  <span style='background:#EFF6FF;color:#1D4ED8;padding:3px 12px;border-radius:20px;font-size:12px;font-weight:700;border:1px solid #93C5FD;'>
                    📊 Avg {_avg:.1f}%
                  </span>
                </div>""", unsafe_allow_html=True)
        elif status_filter == "Suggestions":
            # Exclude Group Match rows — those have their own dedicated filter
            df_view = result_display[
                result_display['Recon_Status'].str.contains('Suggestion', na=False) &
                ~result_display['Recon_Status'].str.contains('Group Match', na=False)
            ].copy()
            if 'GSTIN_BOOKS' in df_view.columns and 'GSTIN_GST' in df_view.columns:
                df_view.insert(0, 'GSTIN Match?',
                               np.where(df_view['GSTIN_BOOKS'] == df_view['GSTIN_GST'], '✅ Same', '❌ Different'))
        elif status_filter == "🔗 Group Match":
            df_view = result_display[
                result_display['Recon_Status'].str.contains('Group Match', na=False)
            ].copy()
            # Identify which side of the match each row is
            has_books = df_view.get('Taxable Value_BOOKS', pd.Series(dtype=float)).notna()
            has_gst   = df_view.get('Taxable Value_GST', pd.Series(dtype=float)).notna()
            df_view.insert(0, 'Side',
                np.where(has_books & ~has_gst, '📚 Books', np.where(~has_books & has_gst, '🏛️ Portal', '↔ Both')))
            if len(df_view) > 0:
                _gm_gstins = df_view['GSTIN'].dropna().unique()
                st.info(f"🔗 **Group Match** — {len(df_view)} invoice row(s) across **{len(_gm_gstins)} GSTIN(s)** where total values match by GSTIN. These are paired suggestions — both Books and Portal sides shown separately.")
        elif status_filter == "Manually Linked":
            df_view = result_display[result_display['Recon_Status'].str.contains('Manual', na=False)]
        elif status_filter == "Not in 2B":
            df_view = result_display[result_display['Recon_Status'] == "Invoices Not in GSTR-2B"]
        elif status_filter == "Not in Books":
            df_view = result_display[result_display['Recon_Status'] == "Invoices Not in Purchase Books"]

        # ── Live search filter ────────────────────────────────────────────────
        if _live_search:
            _q = _live_search.lower()
            _mask = pd.Series(False, index=df_view.index)
            for _col in ['Name of Party', 'Invoice Number_BOOKS', 'Invoice Number_GST', 'GSTIN_BOOKS', 'GSTIN_GST', 'GSTIN']:
                if _col in df_view.columns:
                    _mask |= df_view[_col].astype(str).str.lower().str.contains(_q, na=False)
            df_view = df_view[_mask]
            if df_view.empty:
                st.warning(f"No rows match **'{_live_search}'**.")
            else:
                st.caption(f"🔍 {len(df_view)} row(s) matching **'{_live_search}'**")

        # Status legend
        st.markdown(
            """<div style='display:flex;flex-wrap:wrap;gap:6px;margin-bottom:8px;'>
            <span style='background:#FFF2F2;color:#C00000;padding:3px 10px;border-radius:12px;font-size:12px;font-weight:600;border:1px solid #C00000'>🔴 Not in 2B</span>
            <span style='background:#FFFBEA;color:#B8860B;padding:3px 10px;border-radius:12px;font-size:12px;font-weight:600;border:1px solid #B8860B'>🟡 Not in Books</span>
            <span style='background:#FFF0F0;color:#C00000;padding:3px 10px;border-radius:12px;font-size:12px;font-weight:600;border:1px solid #C00000'>🔴 Value Mismatch</span>
            <span style='background:#EBF3FB;color:#2E75B6;padding:3px 10px;border-radius:12px;font-size:12px;font-weight:600;border:1px solid #2E75B6'>🔵 Date/Inv Mismatch</span>
            <span style='background:#F0FFF4;color:#1E6B3C;padding:3px 10px;border-radius:12px;font-size:12px;font-weight:600;border:1px solid #1E6B3C'>🟢 Matched</span>
            <span style='background:#FDF4FF;color:#7C3AED;padding:3px 10px;border-radius:12px;font-size:12px;font-weight:600;border:1px solid #7C3AED'>🔗 Group Match</span>
            </div>""", unsafe_allow_html=True
        )
                # Status-based row coloring
        STATUS_COLORS_DF = {
            'Invoices Not in GSTR-2B':       'background-color:#FFF2F2; color:#C00000; font-weight:600',
            'Invoices Not in Purchase Books': 'background-color:#FFFBEA; color:#B8860B; font-weight:600',
            'AI Matched (Mismatch)':          'background-color:#FFF0F0; color:#C00000',
            'Matched (Tax Error)':            'background-color:#FFFBEA; color:#B8860B',
            'AI Matched (Date Mismatch)':     'background-color:#EBF3FB; color:#2E75B6',
            'AI Matched (Invoice Mismatch)':  'background-color:#EBF3FB; color:#2E75B6',
            'Matched':                        'background-color:#F0FFF4; color:#1E6B3C',
            'Suggestion (Group Match)':       'background-color:#FDF4FF; color:#7C3AED; font-weight:600',
            'Suggestion':                     'background-color:#EFF4FF; color:#2E75B6',
            'Manually Linked':                'background-color:#F0FFF4; color:#1E6B3C',
        }
        def _color_rows(row):
            st_val = str(row.get('Recon_Status', ''))
            for key, style in STATUS_COLORS_DF.items():
                if key in st_val:
                    return [style] * len(row)
            return [''] * len(row)

        if 'Recon_Status' in df_view.columns and len(df_view) < 5000:
            styled_df = df_view.style.apply(_color_rows, axis=1)
            st.dataframe(
                styled_df, use_container_width=True,
                column_config={
                    "Recon_Status":        st.column_config.TextColumn("Status", width="medium"),
                    "Match_Confidence":    st.column_config.ProgressColumn("AI Confidence %", format="%.1f%%", min_value=0, max_value=100, width="small"),
                    "Taxable Value_BOOKS": st.column_config.NumberColumn("Books Taxable", format="₹ %.2f"),
                    "Taxable Value_GST":   st.column_config.NumberColumn("Portal Taxable", format="₹ %.2f"),
                    "Final_Taxable":       st.column_config.NumberColumn("Final Taxable", format="₹ %.2f"),
                    "IGST_BOOKS":          st.column_config.NumberColumn("IGST (Books)", format="₹ %.2f"),
                    "CGST_BOOKS":          st.column_config.NumberColumn("CGST (Books)", format="₹ %.2f"),
                    "SGST_BOOKS":          st.column_config.NumberColumn("SGST (Books)", format="₹ %.2f"),
                    "IGST_GST":            st.column_config.NumberColumn("IGST (Portal)", format="₹ %.2f"),
                    "CGST_GST":            st.column_config.NumberColumn("CGST (Portal)", format="₹ %.2f"),
                    "SGST_GST":            st.column_config.NumberColumn("SGST (Portal)", format="₹ %.2f"),
                }
            )
        else:
            st.dataframe(
                df_view, use_container_width=True,
                column_config={
                    "Recon_Status":        st.column_config.TextColumn("Status", width="medium"),
                    "Taxable Value_BOOKS": st.column_config.NumberColumn("Books Taxable", format="₹ %.2f"),
                    "Taxable Value_GST":   st.column_config.NumberColumn("Portal Taxable", format="₹ %.2f"),
                    "Final_Taxable":       st.column_config.NumberColumn("Final Taxable", format="₹ %.2f"),
                }
            )

    # ─────────────────────────────────────────────────────
    # TAB 4 — SUPPLIER WISE
    # ─────────────────────────────────────────────────────
    with tab5:
        pivot = result.groupby('Name of Party').agg(
            Total_Invoices  =('GSTIN_BOOKS' if 'GSTIN_BOOKS' in result.columns else 'GSTIN', 'count'),
            Taxable_Value   =('Final_Taxable', 'sum'),
            Unmatched_Count =('Recon_Status', lambda x: x.str.contains('Not in', na=False).sum())
        ).reset_index().sort_values('Unmatched_Count', ascending=False)
        st.dataframe(
            pivot, use_container_width=True, hide_index=True,
            column_config={
                "Name of Party":  "Vendor",
                "Total_Invoices": st.column_config.NumberColumn("Total Inv"),
                "Taxable_Value":  st.column_config.NumberColumn("Total Business (₹)", format="₹ %.2f"),
                "Unmatched_Count": st.column_config.NumberColumn("Discrepancies")
            }
        )

    # ─────────────────────────────────────────────────────
    # TAB 5 — MANUAL MATCHER
    # ─────────────────────────────────────────────────────
    with tab6:
        c1, c2 = st.columns([2, 1])
        with c1: st.write("🔗 **Link Unmatched Invoices Manually**")
        with c2:
            if st.button("Clear All Manual Links", type="secondary"):
                st.session_state.manual_matches = []
                st.session_state.app_stage = 'processing'
                st.rerun()

        unmatched_books = result[result['Recon_Status'] == "Invoices Not in GSTR-2B"].copy()
        unmatched_gst   = result[result['Recon_Status'] == "Invoices Not in Purchase Books"].copy()

        unmatched_books['Label'] = unmatched_books.apply(
            lambda x: f"{x['Name of Party']} | Inv: {x.get('Invoice Number_BOOKS','')} | ₹{x.get('Taxable Value_BOOKS',0)}", axis=1)
        unmatched_gst['Label']   = unmatched_gst.apply(
            lambda x: f"{x['Name of Party']} | Inv: {x.get('Invoice Number_GST','')} | ₹{x.get('Taxable Value_GST',0)}", axis=1)

        if 'Unique_ID_BOOKS' in unmatched_books.columns: unmatched_books['ID'] = unmatched_books['Unique_ID_BOOKS']
        elif 'Unique_ID' in unmatched_books.columns:     unmatched_books['ID'] = unmatched_books['Unique_ID']
        if 'Unique_ID_GST' in unmatched_gst.columns:    unmatched_gst['ID']   = unmatched_gst['Unique_ID_GST']
        elif 'Unique_ID' in unmatched_gst.columns:      unmatched_gst['ID']   = unmatched_gst['Unique_ID']

        col_left, col_mid, col_right = st.columns([1, 0.2, 1])
        with col_left:  b_choice = st.selectbox("Select Invoice from Books",    unmatched_books['Label'].tolist(), index=None)
        with col_mid:   st.markdown("<h2 style='text-align:center;color:#aaa;'>🔗</h2>", unsafe_allow_html=True)
        with col_right: g_choice = st.selectbox("Select Invoice from GSTR-2B", unmatched_gst['Label'].tolist(),   index=None)

        if st.button("Link Selected Pair", type="primary", use_container_width=True):
            if b_choice and g_choice:
                b_id = unmatched_books[unmatched_books['Label'] == b_choice]['ID'].iloc[0]
                g_id = unmatched_gst[unmatched_gst['Label']     == g_choice]['ID'].iloc[0]
                st.session_state.manual_matches.append((b_id, g_id))
                if st.session_state.current_recon_id:
                    log_action(st.session_state.current_recon_id, 'manual_link',
                               {'books_id': str(b_id), 'gst_id': str(g_id),
                                'books_label': b_choice[:80], 'gst_label': g_choice[:80]})
                st.success("Linked! Re-running reconciliation...")
                st.session_state.app_stage = 'processing'
                st.rerun()

        if st.session_state.manual_matches:
            st.info(f"{len(st.session_state.manual_matches)} manual link(s) active.")

    # ─────────────────────────────────────────────────────
    # TAB 6 — VENDOR COMMS
    # ─────────────────────────────────────────────────────
    with tab7:
        st.subheader("💬 Vendor Communication Center")

        # ─────────────────────────────────────────────────────────────────────
        # IMPORT RECONCILED EXCEL → GENERATE NOTICES
        # ─────────────────────────────────────────────────────────────────────
        with st.expander("📥 Import Reconciled Excel → Generate Notices & Action Report", expanded=False):
            st.markdown("""
            <div style='background:var(--navy);color:white;padding:12px 18px;border-radius:var(--r-xs);margin-bottom:12px;'>
                <b>📥 Upload your final reconciliation result Excel</b><br>
                <span style='font-size:12px;opacity:0.85;'>
                The tool auto-detects columns. Supports the tool's own exported format
                and manually arranged files. Required headers: <b>GSTIN, Name of Party,
                Status</b> (partial names accepted).
                </span>
            </div>
            """, unsafe_allow_html=True)

            _tmpl_buf = io.BytesIO()
            _tmpl_sample = pd.DataFrame([{
                'GSTIN': '24ABCDE1234F1Z5', 'Name of Party': 'SAMPLE VENDOR',
                'Invoice Number': 'INV-001', 'Invoice Date': '01/04/2025',
                'Taxable Value': 10000, 'IGST': 1800, 'CGST': 0, 'SGST': 0,
                'Recon_Status': 'Invoices Not in GSTR-2B',
            }])
            with pd.ExcelWriter(_tmpl_buf, engine='xlsxwriter') as _tw:
                _tmpl_sample.to_excel(_tw, sheet_name='Reconciliation', index=False)
                _wb = _tw.book; _ws = _tw.sheets['Reconciliation']
                _hfmt = _wb.add_format({'bold': True, 'bg_color': '#1B2035', 'font_color': 'white', 'border': 1})
                for _ci, _ch in enumerate(_tmpl_sample.columns):
                    _ws.write(0, _ci, _ch, _hfmt); _ws.set_column(_ci, _ci, 22)
            _tmpl_buf.seek(0)
            st.download_button("📋 Download Import Template (.xlsx)", data=_tmpl_buf.getvalue(),
                               file_name="GST_Recon_Import_Template.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            st.caption("Required columns: GSTIN, Name of Party, Recon_Status. Optional: Invoice Number, Taxable Value, IGST, CGST, SGST.")
            st.markdown("---")

            imp_file = st.file_uploader("Upload Result Excel (.xlsx)", type=["xlsx"],
                                        key="notice_import_uploader",
                                        help="Upload the reconciliation result file. Header row is auto-detected.")
            if imp_file:
                imp_bytes = imp_file.read()
                imp_sheets = get_available_sheets(imp_bytes)
                imp_sheet = None
                if len(imp_sheets) > 1:
                    imp_sheet = st.selectbox("Select Sheet", imp_sheets, key="imp_sheet_sel")
                with st.spinner("Detecting columns and importing data…"):
                    imp_df, col_map, missing_cols, imp_warnings = parse_uploaded_result_excel(imp_bytes, sheet_name=imp_sheet)
                if imp_warnings:
                    for w in imp_warnings: st.caption(f"ℹ️ {w}")
                if missing_cols:
                    st.error(f"❌ Required columns not found: **{', '.join(missing_cols)}**  \nPlease ensure your file has columns for GSTIN, Name of Party, and Status.")
                elif imp_df is None or imp_df.empty:
                    st.warning("⚠️ No data found in the file.")
                else:
                    st.success(f"✅ Imported **{len(imp_df):,} rows** from '{imp_file.name}'  |  **{len(col_map)}** columns mapped automatically.")
                    with st.expander("🗂️ Column Mapping Details", expanded=False):
                        st.dataframe(pd.DataFrame([{"Original Column": r, "→ Internal Name": i} for r, i in col_map.items()]), hide_index=True)

                    imp_issue_mask = imp_df['Recon_Status'].str.contains('Not in|Mismatch|Suggestion|Manual|Tax Error', na=False)
                    imp_vendors = sorted([v for v in imp_df[imp_issue_mask]['Name of Party'].dropna().astype(str).unique() if v and v != 'nan'])

                    if not imp_vendors:
                        st.info("No discrepancy vendors found. Check the 'Status' column values in your file.")
                    else:
                        st.markdown(f"**{len(imp_vendors)} vendor(s) with issues found.**")
                        st_counts = imp_df[imp_issue_mask]['Recon_Status'].value_counts()
                        if not st_counts.empty:
                            _cols_st = st.columns(min(len(st_counts), 4))
                            for i, (st_name, cnt) in enumerate(st_counts.items()):
                                with _cols_st[i % 4]:
                                    st.metric(st_name.replace("Invoices ","").replace("AI Matched ","")[:28], cnt)
                        st.markdown("---")
                        _imp_c1, _imp_c2 = st.columns([3, 1])
                        with _imp_c1:
                            imp_selected_vendors = st.multiselect("Select vendors to generate notices:", imp_vendors,
                                default=imp_vendors[:5] if len(imp_vendors) <= 10 else [], key="imp_vendor_sel")
                        with _imp_c2:
                            imp_company = st.text_input("Your Company Name", value=st.session_state.get('meta_name',''), key="imp_company_name")

                        if imp_selected_vendors and imp_company:
                            _imp_pdf_col, _imp_wa_col = st.columns(2)
                            with _imp_pdf_col:
                                if st.button("📄 Generate PDF Notices", type="primary", use_container_width=True, key="imp_pdf_btn"):
                                    imp_zip_buf = io.BytesIO(); imp_errors = []
                                    with zipfile.ZipFile(imp_zip_buf, "a", zipfile.ZIP_DEFLATED, False) as imp_zip:
                                        for v in imp_selected_vendors:
                                            try:
                                                imp_gstin = str(imp_df[imp_df['Name of Party'] == v]['GSTIN'].iloc[0]) if 'GSTIN' in imp_df.columns else ''
                                                imp_zip.writestr(f"GST_Notice_{v}.pdf", create_vendor_pdf(imp_df, v, imp_company, imp_gstin).getvalue())
                                            except Exception as _pe: imp_errors.append(f"{v}: {_pe}")
                                    if imp_errors: st.warning("Some PDFs failed: " + "; ".join(imp_errors))
                                    st.download_button("⬇️ Download PDF Notices ZIP", data=imp_zip_buf.getvalue(),
                                        file_name=f"GST_Notices_Import_{pd.Timestamp.now().strftime('%Y%m%d')}.zip",
                                        mime="application/zip", use_container_width=True, key="imp_pdf_dl")
                            with _imp_wa_col:
                                imp_wa_vendor = st.selectbox("Preview WhatsApp for:", imp_selected_vendors, key="imp_wa_vendor")
                                if st.button("📱 Preview WhatsApp Message", use_container_width=True, key="imp_wa_btn"):
                                    st.session_state['imp_wa_preview'] = generate_whatsapp_message(imp_df, imp_wa_vendor, imp_company) or "No issues found."
                            if st.session_state.get('imp_wa_preview'):
                                st.code(st.session_state['imp_wa_preview'], language='markdown')
                            st.markdown("---")
                            st.markdown("#### 📑 Action Report PDF")
                            _imp_fy = st.text_input("Financial Year", value=st.session_state.get('meta_fy','2025 - 2026'), key="imp_rpt_fy")
                            _imp_per = st.text_input("Period", value=st.session_state.get('meta_period',''), key="imp_rpt_per")
                            _imp_gst = st.text_input("GSTIN", value=st.session_state.get('meta_gstin',''), key="imp_rpt_gstin")
                            if st.button("📑 Generate Action Report PDF", type="primary", use_container_width=True, key="imp_action_rpt_btn"):
                                with st.spinner("Generating…"):
                                    try:
                                        _action_pdf = create_action_report_pdf(imp_df[imp_df['Name of Party'].isin(imp_selected_vendors)], imp_company, _imp_gst, _imp_per, _imp_fy)
                                        st.download_button("⬇️ Download Action Report PDF", data=_action_pdf.getvalue(),
                                            file_name=f"GST_Action_Report_{_imp_per}.pdf", mime="application/pdf",
                                            type="primary", use_container_width=True, key="imp_action_rpt_dl")
                                    except Exception as _are: st.error(f"Report error: {_are}")
                        elif imp_selected_vendors and not imp_company:
                            st.warning("Please enter your company name above.")

        with st.expander("🌐 Fix 'Unknown' Vendors — Enter Name Manually", expanded=False):
            st.info("💡 If any vendor appears as 'Unknown', look up their name on the [GST Portal](https://www.gst.gov.in/searchtaxpayer) and enter it below.")
            if 'Name of Party' in result.columns:
                unique_unknown_gstins = result[result['Name of Party'] == 'Unknown']['GSTIN'].dropna().unique().tolist()
                unique_unknown_gstins = [g for g in unique_unknown_gstins if g and str(g) not in ('', 'nan')]
            else:
                unique_unknown_gstins = []
            if len(unique_unknown_gstins) > 0:
                st.warning(f"Found {len(unique_unknown_gstins)} GSTIN(s) with name 'Unknown'.")
                _unk_cols = st.columns([2, 3, 1])
                with _unk_cols[0]: _unk_gstin_sel = st.selectbox("Select GSTIN", unique_unknown_gstins, key="unk_gstin_sel")
                with _unk_cols[1]: _unk_name_inp = st.text_input("Correct Vendor Name", placeholder="e.g. ACME TEXTILES PVT LTD", key="unk_name_inp")
                with _unk_cols[2]:
                    st.markdown("<br>", unsafe_allow_html=True)
                    if st.button("✅ Apply", key="unk_apply_btn", use_container_width=True):
                        if _unk_name_inp.strip():
                            st.session_state['last_result'].loc[st.session_state['last_result']['GSTIN'] == _unk_gstin_sel, 'Name of Party'] = _unk_name_inp.strip()
                            if st.session_state.current_recon_id:
                                log_action(st.session_state.current_recon_id, 'name_change', {'gstin': _unk_gstin_sel, 'new_name': _unk_name_inp.strip()})
                                save_reconciliation({'gstin': gstin, 'name': name, 'fy': fy, 'period': period}, st.session_state['last_result'])
                            st.success(f"✅ Updated '{_unk_gstin_sel}' → {_unk_name_inp.strip()}")
                            st.rerun()
                        else: st.warning("Please enter a vendor name first.")
            else:
                st.success("✅ All vendors identified!")

        with st.expander("✎ Correct Vendor Name (Fix 'Unknown' by GSTIN)", expanded=False):
            _b2b_parties = result['Name of Party'].dropna().astype(str).unique().tolist()
            _cdnr_result = st.session_state.get('cdnr_result')
            if _cdnr_result is not None and not _cdnr_result.empty and 'Name of Party' in _cdnr_result.columns:
                all_parties = sorted(set(_b2b_parties) | set(_cdnr_result['Name of Party'].dropna().astype(str).unique().tolist()))
            else:
                all_parties = sorted(_b2b_parties)
            default_ix    = all_parties.index('Unknown') if 'Unknown' in all_parties else 0
            target_vendor = st.selectbox("Select Vendor Name to Fix", all_parties, index=default_ix, key="fix_target_name")
            _b2b_gstins = result[result['Name of Party'] == target_vendor][['GSTIN']]
            _gstin_rows = [_b2b_gstins]
            if _cdnr_result is not None and not _cdnr_result.empty:
                _gstin_col = 'GSTIN_BOOKS' if 'GSTIN_BOOKS' in _cdnr_result.columns else 'GSTIN'
                if 'Name of Party' in _cdnr_result.columns:
                    _gstin_rows.append(_cdnr_result[_cdnr_result['Name of Party'] == target_vendor][[_gstin_col]].rename(columns={_gstin_col:'GSTIN'}))
            df_renames = pd.concat(_gstin_rows, ignore_index=True).drop_duplicates(subset=['GSTIN']).reset_index(drop=True)
            df_renames['New Name'] = target_vendor
            _src_note = " (from B2B + CDNR)" if _cdnr_result is not None else " (from B2B)"
            st.markdown(f"**Update Names for '{target_vendor}'{_src_note}:** (Edit the 'New Name' column below)")
            edited_renames = st.data_editor(df_renames, use_container_width=True, hide_index=True,
                column_config={"GSTIN": st.column_config.TextColumn("GSTIN", disabled=True),
                               "New Name": st.column_config.TextColumn("New Name (editable)")})
            if st.button("✅ Apply Name Changes", type="primary", key="apply_name_changes"):
                changes_count = 0
                for _, rr in edited_renames.iterrows():
                    new_n = str(rr['New Name']).strip()
                    _g = str(rr['GSTIN']).strip()
                    if new_n and new_n != target_vendor:
                        _m = st.session_state['last_result']['GSTIN'] == _g
                        st.session_state['last_result'].loc[_m, 'Name of Party'] = new_n
                        if _cdnr_result is not None and not _cdnr_result.empty:
                            _gc = 'GSTIN_BOOKS' if 'GSTIN_BOOKS' in _cdnr_result.columns else 'GSTIN'
                            _mc = st.session_state['cdnr_result'][_gc] == _g
                            st.session_state['cdnr_result'].loc[_mc, 'Name of Party'] = new_n
                        if st.session_state.current_recon_id:
                            log_action(st.session_state.current_recon_id, 'name_change', {'gstin': _g, 'new_name': new_n})
                        changes_count += 1
                if changes_count:
                    if st.session_state.current_recon_id:
                        save_reconciliation({'gstin': gstin, 'name': name, 'fy': fy, 'period': period}, st.session_state['last_result'])
                    st.success(f"✅ Updated {changes_count} GSTIN(s) successfully!")
                    time.sleep(1); st.rerun()
                else:
                    st.info("No changes detected.")

        result = st.session_state['last_result']
        issue_vendors = get_vendors_with_issues(result)

        # ════════════════════════════════════════════════════════════════════
        # ── SECTION 1: CATEGORY-EXCLUSIVE NOTICES (NEW) ──────────────────
        # ════════════════════════════════════════════════════════════════════
        st.markdown("---")
        st.markdown("### 📤 Send Targeted Notices")

        # Language selector — prominent, applies to ALL notice types
        st.markdown("""
        <div style="background:var(--amber-lt);border-radius:var(--r-xs);padding:10px 16px;
                    display:flex;align-items:center;gap:10px;margin-bottom:12px">
          <span style="font-size:16px">🌐</span>
          <span style="font-size:12px;font-weight:700;color:var(--amber-dk)">
            Choose notice language — applies to WhatsApp messages and bulk .txt export
          </span>
        </div>
        """, unsafe_allow_html=True)
        _lang_sel = st.radio("Language", ["🇬🇧 English", "🇮🇳 Hindi", "🇮🇳 Gujarati"],
                             horizontal=True, key="global_lang_radio")
        _global_lang = 'en' if 'English' in _lang_sel else ('hi' if 'Hindi' in _lang_sel else 'gu')
        st.session_state['wa_lang'] = _global_lang

        # ── TWO EXCLUSIVE CATEGORY PANELS ────────────────────────────────────
        _cat_col1, _cat_col2 = st.columns(2, gap="medium")

        # Category counts
        _not2b_vendors   = get_vendors_by_category(result, 'not_in_2b')
        _notbooks_vendors = get_vendors_by_category(result, 'not_in_books')

        with _cat_col1:
            st.markdown(f"""
            <div style="background:var(--red-lt);border-radius:var(--r-sm);padding:14px 16px;margin-bottom:10px">
              <div style="font-size:13px;font-weight:800;color:var(--red)">🔴 Not in GSTR-2B</div>
              <div style="font-size:11px;color:#A82C18;margin-top:3px">
                Your invoice is in our books but NOT in your GSTR-2B portal filing.
                Vendor must upload in GSTR-1.
              </div>
              <div style="font-size:20px;font-weight:800;color:var(--red);margin-top:8px">{len(_not2b_vendors)}</div>
              <div style="font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:#A82C18">vendors affected</div>
            </div>
            """, unsafe_allow_html=True)

            if _not2b_vendors:
                _sel_2b = st.multiselect("Select vendors (Not in 2B):", _not2b_vendors,
                                         default=_not2b_vendors, key="cat_2b_vendors")
                if _sel_2b:
                    _c2b_a, _c2b_b = st.columns(2)
                    with _c2b_a:
                        # Preview single
                        _prev_2b = st.selectbox("Preview notice for:", _sel_2b, key="prev_2b_vendor")
                        if st.button("👁 Preview Notice", key="prev_2b_btn", use_container_width=True):
                            _msg = generate_targeted_notice(result, _prev_2b, name, 'not_in_2b', _global_lang)
                            st.session_state['preview_2b_msg'] = _msg
                    with _c2b_b:
                        # Bulk WA txt
                        _wa_2b_lines = []
                        for _v2b in _sel_2b:
                            _m = generate_targeted_notice(result, _v2b, name, 'not_in_2b', _global_lang)
                            _wa_2b_lines.append(f"{'='*50}\nVENDOR: {_v2b}\n{'='*50}\n{_m}\n\n")
                        _wa_2b_bytes = "\n".join(_wa_2b_lines).encode('utf-8')
                        st.download_button("📱 Bulk WA — Not in 2B (.txt)", data=_wa_2b_bytes,
                                           file_name=f"NotIn2B_Notices_{period}.txt",
                                           mime="text/plain", use_container_width=True,
                                           key="bulk_2b_wa_dl",
                                           help="One file, one section per vendor. Copy-paste each into WhatsApp.")
                    if st.session_state.get('preview_2b_msg'):
                        st.markdown("**📋 Notice Preview:**")
                        st.code(st.session_state['preview_2b_msg'], language='markdown')
                    # PDF ZIP for Not in 2B
                    if st.button("📄 PDF Notices ZIP — Not in 2B", type="primary",
                                 use_container_width=True, key="pdf_2b_btn"):
                        _z2b = io.BytesIO()
                        with zipfile.ZipFile(_z2b, "a", zipfile.ZIP_DEFLATED, False) as _zf:
                            for _v in _sel_2b:
                                _pdf = create_vendor_pdf(result[result['Recon_Status'] == 'Invoices Not in GSTR-2B'], _v, name, gstin)
                                _zf.writestr(f"NotIn2B_Notice_{_v}.pdf", _pdf.getvalue())
                        st.download_button("⬇️ Download ZIP", data=_z2b.getvalue(),
                                           file_name=f"NotIn2B_Notices_{period}.zip",
                                           mime="application/zip", use_container_width=True, key="pdf_2b_dl")
            else:
                st.success("✅ No vendors with 'Not in 2B' invoices.")

        with _cat_col2:
            st.markdown(f"""
            <div style="background:var(--amber-lt);border-radius:var(--r-sm);padding:14px 16px;margin-bottom:10px">
              <div style="font-size:13px;font-weight:800;color:var(--amber-dk)">🟠 Not in Our Books</div>
              <div style="font-size:11px;color:var(--amber-md);margin-top:3px">
                Invoice appears in GST portal / GSTR-2B but is NOT in our Purchase Register.
                Vendor must send invoice copy or issue Credit Note.
              </div>
              <div style="font-size:20px;font-weight:800;color:var(--amber-dk);margin-top:8px">{len(_notbooks_vendors)}</div>
              <div style="font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:var(--amber-md)">vendors affected</div>
            </div>
            """, unsafe_allow_html=True)

            if _notbooks_vendors:
                _sel_nb = st.multiselect("Select vendors (Not in Books):", _notbooks_vendors,
                                         default=_notbooks_vendors, key="cat_nb_vendors")
                if _sel_nb:
                    _cnb_a, _cnb_b = st.columns(2)
                    with _cnb_a:
                        _prev_nb = st.selectbox("Preview notice for:", _sel_nb, key="prev_nb_vendor")
                        if st.button("👁 Preview Notice", key="prev_nb_btn", use_container_width=True):
                            _msg_nb = generate_targeted_notice(result, _prev_nb, name, 'not_in_books', _global_lang)
                            st.session_state['preview_nb_msg'] = _msg_nb
                    with _cnb_b:
                        _wa_nb_lines = []
                        for _vnb in _sel_nb:
                            _m = generate_targeted_notice(result, _vnb, name, 'not_in_books', _global_lang)
                            _wa_nb_lines.append(f"{'='*50}\nVENDOR: {_vnb}\n{'='*50}\n{_m}\n\n")
                        _wa_nb_bytes = "\n".join(_wa_nb_lines).encode('utf-8')
                        st.download_button("📱 Bulk WA — Not in Books (.txt)", data=_wa_nb_bytes,
                                           file_name=f"NotInBooks_Notices_{period}.txt",
                                           mime="text/plain", use_container_width=True,
                                           key="bulk_nb_wa_dl")
                    if st.session_state.get('preview_nb_msg'):
                        st.markdown("**📋 Notice Preview:**")
                        st.code(st.session_state['preview_nb_msg'], language='markdown')
                    if st.button("📄 PDF Notices ZIP — Not in Books", type="primary",
                                 use_container_width=True, key="pdf_nb_btn"):
                        _znb = io.BytesIO()
                        with zipfile.ZipFile(_znb, "a", zipfile.ZIP_DEFLATED, False) as _zf:
                            for _v in _sel_nb:
                                _pdf = create_vendor_pdf(result[result['Recon_Status'] == 'Invoices Not in Purchase Books'], _v, name, gstin)
                                _zf.writestr(f"NotInBooks_Notice_{_v}.pdf", _pdf.getvalue())
                        st.download_button("⬇️ Download ZIP", data=_znb.getvalue(),
                                           file_name=f"NotInBooks_Notices_{period}.zip",
                                           mime="application/zip", use_container_width=True, key="pdf_nb_dl")
            else:
                st.success("✅ No vendors with 'Not in Books' invoices.")

        # ════════════════════════════════════════════════════════════════════
        # ── SECTION 2: BULK NOTICE (ALL ISSUES) ──────────────────────────
        # ════════════════════════════════════════════════════════════════════
        st.markdown("---")
        st.markdown("### 📦 Bulk Notice — All Issue Types")

        STATUS_FILTER_OPTS = {
            "All Issues":                  None,
            "Not in GSTR-2B":              "Invoices Not in GSTR-2B",
            "Not in Books":                "Invoices Not in Purchase Books",
            "Value Mismatch":              "AI Matched (Mismatch)",
            "Tax Error":                   "Matched (Tax Error)",
            "Date Mismatch":               "AI Matched (Date Mismatch)",
            "Invoice No. Mismatch":        "AI Matched (Invoice Mismatch)",
            "Suggestions":                 "Suggestion",
        }

        if issue_vendors:
            _c1, _c2 = st.columns([2, 2])
            with _c1:
                bulk_status_filter = st.selectbox("Filter vendors by issue type:", list(STATUS_FILTER_OPTS.keys()), index=0, key="bulk_status_filter")
            selected_status_key = STATUS_FILTER_OPTS[bulk_status_filter]
            if selected_status_key:
                filtered_vendors = [v for v in result[result['Recon_Status'].str.contains(
                    selected_status_key.replace('(','\\(').replace(')','\\)'), na=False)]['Name of Party'].unique()
                    if v and str(v) != 'nan']
            else:
                filtered_vendors = issue_vendors
            with _c2:
                st.metric("Vendors with this issue", len(filtered_vendors))

            selected_vendors_bulk = st.multiselect(
                f"Select vendors ({len(filtered_vendors)} available):",
                filtered_vendors if selected_status_key else issue_vendors,
                default=filtered_vendors if selected_status_key else [],
                key="bulk_vendor_select"
            )

            if selected_vendors_bulk:
                sel_df = result[result['Name of Party'].isin(selected_vendors_bulk)]
                st_counts = sel_df[sel_df['Recon_Status'].str.contains('Not in|Mismatch|Suggestion|Manual|Tax Error', na=False)]['Recon_Status'].value_counts()
                if not st_counts.empty:
                    cols_st = st.columns(min(len(st_counts), 4))
                    for i, (st_name, cnt) in enumerate(st_counts.items()):
                        with cols_st[i % 4]:
                            st.metric(st_name.replace("Invoices ","").replace("AI Matched ",""), cnt)

                c_pdf, c_xls, c_watxt = st.columns(3)
                zip_buffer_pdf = io.BytesIO()
                with zipfile.ZipFile(zip_buffer_pdf, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
                    for v in selected_vendors_bulk:
                        pdf_data = create_vendor_pdf(result, v, name, gstin)
                        zip_file.writestr(f"GST_Notice_{v}.pdf", pdf_data.getvalue())
                filtered_df    = result[result['Name of Party'].isin(selected_vendors_bulk)]
                zip_buffer_xls = generate_vendor_split_zip(filtered_df)
                folder         = st.session_state.current_client_path

                # Bulk WA .txt uses the global language
                _wa_lines = []
                for v in selected_vendors_bulk:
                    _wmsg = generate_whatsapp_message(result, v, name) if _global_lang == 'en' \
                            else generate_whatsapp_message_multilang(result, v, name, lang=_global_lang)
                    _wa_lines.append(f"{'='*50}\nVENDOR: {v}\n{'='*50}\n{_wmsg}\n\n")
                _wa_txt_bytes = "\n".join(_wa_lines).encode('utf-8')

                with c_pdf:
                    st.download_button("📄 PDF Notices (ZIP)", data=zip_buffer_pdf.getvalue(),
                                       file_name="GST_Notices.zip", type="primary", use_container_width=True,
                                       on_click=save_callback, args=(folder, "GST_Notices.zip", zip_buffer_pdf.getvalue()))
                with c_xls:
                    st.download_button("📊 Excel Splits", data=zip_buffer_xls.getvalue(),
                                       file_name="Excels.zip", use_container_width=True,
                                       on_click=save_callback, args=(folder, "Excels.zip", zip_buffer_xls.getvalue()))
                with c_watxt:
                    st.download_button("📱 Bulk WhatsApp (.txt)", data=_wa_txt_bytes,
                                       file_name=f"WhatsApp_Notices_{period}.txt", mime="text/plain",
                                       use_container_width=True, key="bulk_wa_all_dl")

        # ════════════════════════════════════════════════════════════════════
        # ── SECTION 3: SINGLE VENDOR NOTICE ──────────────────────────────
        # ════════════════════════════════════════════════════════════════════
        st.markdown("---")
        st.markdown("### 📱 Send Notice — Single Vendor")

        if not issue_vendors:
            st.info("No vendors with discrepancies found.")
        else:
            c_vendor, c_mode = st.columns([2, 1])
            with c_vendor:
                selected_vendor = st.selectbox("Select Vendor", issue_vendors, key="single_vendor_sel")
            with c_mode:
                comm_mode = st.radio("Mode", ["📧 Email", "📱 WhatsApp", "📄 Preview PDF"], horizontal=True)

            if selected_vendor:
                v_df = result[(result['Name of Party'] == selected_vendor) &
                               result['Recon_Status'].str.contains('Not in|Mismatch|Suggestion|Manual|Tax Error', na=False)]
                v_counts = v_df['Recon_Status'].value_counts()
                if not v_counts.empty:
                    st.caption("Issues: " + " | ".join(
                        f"**{cnt}×** {st_n.replace('Invoices ','').replace('AI Matched ','')}"
                        for st_n, cnt in v_counts.items()
                    ))

                if comm_mode == "📧 Email":
                    subject, body_txt = generate_email_draft(result, selected_vendor, name)
                    st.text_input("Subject", value=subject, key="email_subj")
                    st.code(body_txt, language='markdown')

                elif comm_mode == "📱 WhatsApp":
                    # Language uses global selector already set above
                    st.markdown(f"**Language:** {_lang_sel}  *(change the language selector above)*")

                    # Also show targeted category options if applicable
                    _v_has_2b = any('Not in GSTR-2B' in s for s in v_df['Recon_Status'].values)
                    _v_has_nb = any('Not in Purchase Books' in s for s in v_df['Recon_Status'].values)

                    _notice_type_opts = ["All Issues (Combined)"]
                    if _v_has_2b:   _notice_type_opts.append("🔴 Not in GSTR-2B Only")
                    if _v_has_nb:   _notice_type_opts.append("🟠 Not in Books Only")
                    _notice_type = st.radio("Notice content:", _notice_type_opts, horizontal=True, key="single_notice_type")

                    if _notice_type == "🔴 Not in GSTR-2B Only":
                        wa_body = generate_targeted_notice(result, selected_vendor, name, 'not_in_2b', _global_lang)
                    elif _notice_type == "🟠 Not in Books Only":
                        wa_body = generate_targeted_notice(result, selected_vendor, name, 'not_in_books', _global_lang)
                    elif _global_lang == 'en':
                        wa_body = generate_whatsapp_message(result, selected_vendor, name)
                    else:
                        wa_body = generate_whatsapp_message_multilang(result, selected_vendor, name, lang=_global_lang)

                    st.code(wa_body, language='markdown')

                    if st.session_state.current_recon_id and st.button("✅ Mark Notice Sent to this Vendor", key="mark_sent_single"):
                        _v_gstin = str(result[result['Name of Party'] == selected_vendor]['GSTIN'].iloc[0]) if 'GSTIN' in result.columns else ''
                        _v_issues = len(v_df)
                        _v_itc    = float(v_df['Final_Taxable'].sum()) if 'Final_Taxable' in v_df.columns else 0.0
                        upsert_followup(st.session_state.current_recon_id, selected_vendor, _v_gstin, _v_issues, _v_itc)
                        save_followup_notice_sent(st.session_state.current_recon_id, selected_vendor)
                        log_action(st.session_state.current_recon_id, 'notice_sent', {'vendor': selected_vendor, 'lang': _global_lang, 'type': _notice_type})
                        st.session_state['notices_sent_count'] = st.session_state.get('notices_sent_count', 0) + 1
                        st.success(f"✅ Notice logged for **{selected_vendor}** — visible in Follow-up Tracker (Tab 8).")
                        st.markdown("""
                        <div class="next-step-hint" style="margin-top:8px">
                          <div style="font-size:11px;font-weight:800;color:#1352C9;letter-spacing:.07em;text-transform:uppercase;margin-bottom:5px">
                            Notice Sent — Next Step
                          </div>
                          <div style="font-size:14px;font-weight:800;color:#0D1B40">
                            <span class="next-step-arrow">→</span> Click <b>📌 Follow-up Tracker (Tab 8)</b> to track responses
                          </div>
                        </div>
                        """, unsafe_allow_html=True)

                elif comm_mode == "📄 Preview PDF":
                    pdf_data = create_vendor_pdf(result, selected_vendor, name, gstin)
                    st.download_button(f"⬇️ Download Notice PDF — {selected_vendor}", data=pdf_data.getvalue(),
                                       file_name=f"GST_Notice_{selected_vendor}.pdf", mime="application/pdf",
                                       type="primary", use_container_width=True, key="single_pdf_dl")

    # ─────────────────────────────────────────────────────
    # TAB 8 — FOLLOW-UP TRACKER
    # ─────────────────────────────────────────────────────
    with tab8:
        _nt_sent = st.session_state.get('notices_sent_count', 0)
        if _nt_sent > 0:
            st.success(f"✅ {_nt_sent} notice(s) sent this session — vendors are tracked below.")
        st.subheader("📌 Vendor Follow-up Tracker")
        st.caption(
            "Tracks vendors with invoices missing from GSTR-2B. "
            "Send notices from Tab 7 → these vendors appear here automatically."
        )

        recon_id_now = st.session_state.current_recon_id

        if not recon_id_now:
            st.info("Open a reconciliation from the sidebar or run one to use the Follow-up Tracker.")
        else:
            # ── Only track "Not in GSTR-2B" — these are the vendors we follow up with
            _FOLLOWUP_MASK = 'Not in GSTR-2B'
            _all_issue_v = sorted(
                result[
                    result['Recon_Status'].str.contains(_FOLLOWUP_MASK, na=False)
                ]['Name of Party'].dropna().unique().tolist()
            )
            _all_issue_v = [v for v in _all_issue_v if v and str(v) != 'nan']

            for _av in _all_issue_v:
                _av_gstin = str(result[result['Name of Party'] == _av]['GSTIN'].iloc[0]) \
                            if 'GSTIN' in result.columns and len(result[result['Name of Party'] == _av]) > 0 else ''
                _av_df    = result[
                    (result['Name of Party'] == _av) &
                    result['Recon_Status'].str.contains(_FOLLOWUP_MASK, na=False)
                ]
                _av_itc   = float(_av_df['Final_Taxable'].sum()) if 'Final_Taxable' in _av_df.columns else 0.0
                upsert_followup(recon_id_now, _av, _av_gstin, len(_av_df), _av_itc)

            followup_df = get_followups(recon_id_now)

            if followup_df.empty:
                st.info("No vendors tracked yet. Run reconciliation and send notices from Tab 7.")
            else:
                # ── Summary KPIs ─────────────────────────────────────────────
                _fp1, _fp2, _fp3, _fp4 = st.columns(4)
                STATUS_COLORS_FU = {
                    'Pending':    '#FFF2F2',
                    'Responded':  '#EBF3FB',
                    'Resolved':   '#F0FFF4',
                    'Escalated':  '#FFF8E1',
                }
                _fp1.metric("Total Vendors",   len(followup_df))
                _fp2.metric("🔴 Pending",        int((followup_df['status'] == 'Pending').sum()))
                _fp3.metric("✅ Resolved",        int((followup_df['status'] == 'Resolved').sum()))
                _fp4.metric("📨 Notice Sent",     int(followup_df['notice_sent_date'].notna().sum()))

                st.markdown("---")

                # ── Inline status editor ─────────────────────────────────────
                st.markdown("#### Update Vendor Follow-up Status")
                st.caption("Select a vendor to update their status and add notes.")

                _fu_vendor = st.selectbox(
                    "Select Vendor",
                    followup_df['vendor_name'].tolist(),
                    key="fu_vendor_sel"
                )
                _fu_row = followup_df[followup_df['vendor_name'] == _fu_vendor].iloc[0]

                _fuc1, _fuc2 = st.columns([1, 2])
                with _fuc1:
                    _new_status = st.selectbox(
                        "Status",
                        ["Pending", "Responded", "Resolved", "Escalated"],
                        index=["Pending", "Responded", "Resolved", "Escalated"].index(
                            _fu_row['status'] if _fu_row['status'] in
                            ["Pending","Responded","Resolved","Escalated"] else "Pending"
                        ),
                        key="fu_status_sel"
                    )
                with _fuc2:
                    _new_notes = st.text_input(
                        "Notes (optional)",
                        value=str(_fu_row['notes']) if pd.notna(_fu_row['notes']) else '',
                        placeholder="e.g. Called vendor, promised to fix by 20th",
                        key="fu_notes_inp"
                    )

                _fu_btn_c1, _fu_btn_c2 = st.columns(2)
                with _fu_btn_c1:
                    if st.button("💾 Save Status", type="primary", use_container_width=True, key="fu_save"):
                        update_followup_status(recon_id_now, _fu_vendor, _new_status, _new_notes)
                        log_action(recon_id_now, 'followup_update',
                                   {'vendor': _fu_vendor, 'status': _new_status})
                        st.success(f"✅ Updated **{_fu_vendor}** → {_new_status}")
                        st.rerun()
                with _fu_btn_c2:
                    if st.button("📨 Mark Notice Sent Today", use_container_width=True, key="fu_mark_sent"):
                        save_followup_notice_sent(recon_id_now, _fu_vendor)
                        log_action(recon_id_now, 'notice_sent', {'vendor': _fu_vendor})
                        st.success(f"📨 Notice date set to today for **{_fu_vendor}**")
                        st.rerun()

                st.markdown("---")
                st.markdown("#### Full Follow-up Register")

                # Display table with color coding
                _display_fu = followup_df[[
                    'vendor_name','gstin','notice_sent_date','status','issue_count','itc_at_risk','notes','last_updated'
                ]].copy()
                _display_fu.columns = [
                    'Vendor','GSTIN','Notice Sent','Status','Issues','ITC at Risk (₹)','Notes','Last Updated'
                ]
                _display_fu['Last Updated'] = _display_fu['Last Updated'].astype(str).str[:16]
                _display_fu['Notice Sent']  = _display_fu['Notice Sent'].fillna('Not Sent')

                st.dataframe(
                    _display_fu, use_container_width=True, hide_index=True,
                    column_config={
                        'Vendor':       st.column_config.TextColumn(width="large"),
                        'GSTIN':        st.column_config.TextColumn(width="medium"),
                        'Notice Sent':  st.column_config.TextColumn(width="small"),
                        'Status':       st.column_config.TextColumn(width="small"),
                        'Issues':       st.column_config.NumberColumn(width="small"),
                        'ITC at Risk (₹)': st.column_config.NumberColumn(format="₹ %.2f", width="medium"),
                        'Notes':        st.column_config.TextColumn(width="large"),
                        'Last Updated': st.column_config.TextColumn(width="medium"),
                    }
                )

                # Download follow-up register as Excel
                _fu_excel_buf = io.BytesIO()
                with pd.ExcelWriter(_fu_excel_buf, engine='xlsxwriter') as _fw:
                    _display_fu.to_excel(_fw, sheet_name='Follow-up Register', index=False)
                _fu_excel_buf.seek(0)
                st.download_button(
                    "📥 Download Follow-up Register (Excel)",
                    data=_fu_excel_buf.getvalue(),
                    file_name=f"Followup_Register_{period}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

    # ─────────────────────────────────────────────────────
    # TAB 2 — CDNR MATCHING
    # ─────────────────────────────────────────────────────
    with tab2:
        st.markdown("### 📋 CDNR Reconciliation — Credit & Debit Notes")
        st.caption(
            "Matches Credit/Debit Notes from your **Books (CDNR sheet)** against **GSTR-2B CDNR sheet**. "
            "Values handled separately from B2B — no pollution of B2B KPIs."
        )

        has_files = (
            st.session_state.get('file_books_bytes') is not None and
            st.session_state.get('file_gst_bytes')   is not None
        )

        if not has_files:
            if st.session_state.cdnr_result is not None:
                st.info("📂 Showing CDNR results loaded from history. Re-upload files to run fresh CDNR.")
            else:
                st.warning("⚠️ Original files not available. Click **🔄 New** and re-upload to run CDNR reconciliation.")
        else:
            if st.button("▶️ Run CDNR Reconciliation", type="primary", use_container_width=True, key="run_cdnr"):
                with st.spinner("Reading CDNR sheets, applying CDNRA amendments, and matching notes..."):
                    try:
                        file_b_io = io.BytesIO(st.session_state['file_books_bytes'])
                        file_g_io = io.BytesIO(st.session_state['file_gst_bytes'])
                        cdnr_result, cdnr_summary = process_cdnr_reconciliation(
                            file_b_io, file_g_io,
                            tolerance  = st.session_state.get('tolerance',   5.0),
                            smart_mode = st.session_state.get('smart_mode', False)
                        )
                        st.session_state.cdnr_result  = cdnr_result
                        st.session_state.cdnr_summary = cdnr_summary
                        # Save to DB so history loads restore CDNR results
                        if st.session_state.current_recon_id:
                            save_cdnr_to_history(st.session_state.current_recon_id, cdnr_result, cdnr_summary)
                            log_action(st.session_state.current_recon_id, 'cdnr_run',
                                       {'matched': cdnr_summary.get('matched_count', 0),
                                        'not_in_2b': cdnr_summary.get('not_in_2b_count', 0)})
                        st.markdown("""
                        <div class="next-step-hint">
                          <div style="font-size:11px;font-weight:800;color:#1352C9;letter-spacing:.07em;text-transform:uppercase;margin-bottom:5px">
                            ✅ CDN Recon Done — Next Step
                          </div>
                          <div style="font-size:14px;font-weight:800;color:#0D1B40">
                            <span class="next-step-arrow">→</span> Click <b>📥 Downloads Hub (Tab 3)</b> to generate & download all reports in one place
                          </div>
                        </div>
                        """, unsafe_allow_html=True)
                    except Exception as e:
                        st.error(f"CDNR Engine Error: {e}")

        if st.session_state.cdnr_result is not None:
            cdnr_result  = st.session_state.cdnr_result
            cdnr_summary = st.session_state.cdnr_summary

            if cdnr_result.empty:
                st.warning(
                    "No CDNR data found. Check that:\n"
                    "- Your Books file has a sheet named exactly **'cdnr'** (lowercase ok)\n"
                    "- Your GSTR-2B file has a **'CDNR'** tab (standard NIC format)"
                )
            else:
                if cdnr_summary.get('amendments_deleted', 0) > 0 or cdnr_summary.get('amendments_added', 0) > 0:
                    st.info(
                        f"⚡ CDNRA Applied: Removed **{cdnr_summary['amendments_deleted']}** old notes, "
                        f"Added **{cdnr_summary['amendments_added']}** revised notes."
                    )

                # KPI Cards (ITC implication section removed)
                k1, k2, k3, k4, k5 = st.columns(5)
                k1.metric("📚 Notes in Books",    cdnr_summary.get('total_books', 0))
                k2.metric("🏛️ Notes in GSTR-2B", cdnr_summary.get('total_gst',   0))
                k3.metric("✅ Matched",            cdnr_summary.get('matched_count', 0))
                k4.metric("⚠️ Value Mismatch",     cdnr_summary.get('mismatch_count', 0))
                k5.metric("❌ Unmatched",
                    cdnr_summary.get('not_in_2b_count', 0) + cdnr_summary.get('not_in_books_count', 0))

                k6, k7 = st.columns(2)
                k6.metric("🔶 Tax Error",  cdnr_summary.get('tax_error_count', 0),
                          help="Taxable matches but IGST/CGST/SGST differs")
                k7.metric("🤖 AI Matched", cdnr_summary.get('ai_matched_count', 0),
                          help="Matched via date/taxable fallback steps")

                st.divider()

                # Filter + CDNR Suggestions tab (shows GSTIN match status like B2B)
                cdnr_filter_opts = [
                    "All Data", "CDNR Matched", "CDNR Matched (Tax Error)",
                    "CDNR AI Matched", "CDNR Mismatch",
                    "CDNR Not in GSTR-2B", "CDNR Not in Books",
                    "⚠️ CDNR Suggestions (Review GSTIN Match)",
                ]
                cdnr_filter = st.selectbox("🔍 Filter CDNR View", cdnr_filter_opts, key="cdnr_filter")

                CDNR_FILTER_MAP = {
                    "CDNR Matched"             : r"CDNR Matched$",
                    "CDNR Matched (Tax Error)" : r"Tax Error",
                    "CDNR AI Matched"          : r"AI Matched",
                    "CDNR Mismatch"            : r"Mismatch",
                    "CDNR Not in GSTR-2B"      : r"Not in GSTR-2B",
                    "CDNR Not in Books"        : r"Not in Books",
                    "⚠️ CDNR Suggestions (Review GSTIN Match)": r"Suggestion",
                }
                df_cdnr_view = cdnr_result.copy()
                if cdnr_filter != "All Data":
                    pat = CDNR_FILTER_MAP.get(cdnr_filter, cdnr_filter)
                    df_cdnr_view = cdnr_result[
                        cdnr_result['Recon_Status_CDNR'].str.contains(pat, regex=True, na=False)
                    ].copy()

                # For Suggestions view: show GSTIN match status column
                if "Suggestion" in cdnr_filter:
                    if 'GSTIN_BOOKS' in df_cdnr_view.columns and 'GSTIN_GST' in df_cdnr_view.columns:
                        df_cdnr_view.insert(0, 'GSTIN Match?',
                            np.where(df_cdnr_view['GSTIN_BOOKS'] == df_cdnr_view['GSTIN_GST'],
                                     '✅ Same GSTIN', '❌ Different GSTIN'))
                    if len(df_cdnr_view) > 0:
                        st.warning(
                            f"⚠️ {len(df_cdnr_view)} CDNR Suggestion(s) found. "
                            "These are cross-GSTIN matches — verify the GSTIN Match column before accepting."
                        )

                st.dataframe(
                    df_cdnr_view, use_container_width=True, hide_index=True,
                    column_config={
                        "Taxable Value_BOOKS": st.column_config.NumberColumn("Taxable (Books)", format="₹ %.2f"),
                        "Taxable Value_GST":   st.column_config.NumberColumn("Taxable (2B)",    format="₹ %.2f"),
                        "Diff_Taxable":        st.column_config.NumberColumn("Diff Taxable",    format="₹ %.2f"),
                        "Diff_IGST":           st.column_config.NumberColumn("Diff IGST",       format="₹ %.2f"),
                        "Diff_CGST":           st.column_config.NumberColumn("Diff CGST",       format="₹ %.2f"),
                        "Diff_SGST":           st.column_config.NumberColumn("Diff SGST",       format="₹ %.2f"),
                        "Recon_Status_CDNR":   st.column_config.TextColumn("Status", width="large"),
                    }
                )

                # Download CDNR Report (quick access — full Downloads Hub in Tab 3)
                try:
                    cdnr_excel_bytes = generate_cdnr_excel(cdnr_result, gstin, name, fy, period,
                                                           b2b_full_df=result)
                except Exception as _err:
                    st.error(f"Report generation error: {_err}")
                    cdnr_excel_bytes = None

                if cdnr_excel_bytes:
                    cdnr_filename = f"CDNR_Reconciliation_{period}.xlsx"
                    folder        = st.session_state.current_client_path
                    st.download_button(
                        label="📥 Download CDNR Reconciliation Report",
                        data=cdnr_excel_bytes,
                        file_name=cdnr_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                        on_click=save_callback,
                        args=(folder, cdnr_filename, cdnr_excel_bytes),
                    )

                # ── UNKNOWN PARTY NAME EDITOR ─────────────────────────────────
                _unknown_mask = cdnr_result['Name of Party'].isin(['Unknown', '', 'nan']) | cdnr_result['Name of Party'].isna()
                _unknown_df = cdnr_result[_unknown_mask][['GSTIN', 'Name of Party']].drop_duplicates(subset=['GSTIN'])
                _unknown_gstins = _unknown_df['GSTIN'].dropna().unique().tolist()
                _unknown_gstins = [g for g in _unknown_gstins if g and str(g) not in ('', 'nan')]

                if _unknown_gstins:
                    with st.expander(f"✏️ Fix Unknown Party Names ({len(_unknown_gstins)} GSTINs)", expanded=True):
                        st.caption("These GSTINs appear in your Books but could not be matched to a supplier name. Enter the correct name manually below.")
                        _name_inputs = {}
                        _cols_per_row = 2
                        _gstin_chunks = [_unknown_gstins[i:i+_cols_per_row] for i in range(0, len(_unknown_gstins), _cols_per_row)]
                        for _chunk in _gstin_chunks:
                            _edit_cols = st.columns(_cols_per_row)
                            for _ci, _g in enumerate(_chunk):
                                with _edit_cols[_ci]:
                                    _name_inputs[_g] = st.text_input(
                                        f"Name for {_g}",
                                        value=st.session_state.get(f'cdnr_name_{_g}', ''),
                                        key=f'cdnr_nameinput_{_g}',
                                        placeholder="Enter supplier name...",
                                    )
                        if st.button("✅ Apply Names to CDNR", type="primary", use_container_width=True, key="apply_cdnr_names"):
                            _updated = 0
                            _new_cdnr = st.session_state.cdnr_result.copy()
                            for _g, _n in _name_inputs.items():
                                _n = _n.strip()
                                if _n:
                                    _mask_g = _new_cdnr['GSTIN'] == _g
                                    _new_cdnr.loc[_mask_g, 'Name of Party'] = _n
                                    st.session_state[f'cdnr_name_{_g}'] = _n
                                    _updated += 1
                            if _updated:
                                st.session_state.cdnr_result = _new_cdnr
                                if st.session_state.get('last_result') is not None:
                                    _last = st.session_state['last_result']
                                    for _g, _n in _name_inputs.items():
                                        _n = _n.strip()
                                        if _n:
                                            _last.loc[(_last['GSTIN'] == _g) & (_last['Name of Party'].isin(['Unknown',''])), 'Name of Party'] = _n
                                    st.session_state['last_result'] = _last
                                st.success(f"✅ Updated {_updated} supplier name(s). Re-generate reports in Downloads Hub (Tab 3) to reflect changes.")
                                st.rerun()
                            else:
                                st.warning("Please enter at least one supplier name before applying.")

                st.divider()
                # ── Point to Downloads Hub ────────────────────────────────────
                st.markdown("""
                <div class="next-step-hint">
                  <div style="font-size:11px;font-weight:800;color:#1352C9;letter-spacing:.07em;text-transform:uppercase;margin-bottom:5px">
                    ✅ CDN Recon Done — Next Step
                  </div>
                  <div style="font-size:14px;font-weight:800;color:#0D1B40">
                    <span class="next-step-arrow">→</span> Click <b>📥 Downloads Hub (Tab 3)</b> to generate the Combined Report, B2B Report &amp; ITC PDF
                  </div>
                </div>
                """, unsafe_allow_html=True)

    # ─────────────────────────────────────────────────────
    # TAB 8 — BACKUP & RESTORE
    # ─────────────────────────────────────────────────────
    with tab9:
        st.markdown("### 💾 Backup & Restore")
        st.markdown(
            "Export your **entire database** (all clients, history, follow-up tracker) as a single file "
            "and restore it on any PC. This lets you continue work — including the Follow-up Tracker — "
            "on a different machine."
        )
        st.markdown("---")

        # ── EXPORT ─────────────────────────────────────────────────────────
        st.markdown("#### 📤 Export Backup")
        st.caption("Downloads the complete `recon_history.db` — contains all reconciliation history, audit logs, and follow-up tracker data.")

        _db_path = "recon_history.db"
        if os.path.exists(_db_path):
            with open(_db_path, "rb") as _dbf:
                _db_bytes = _dbf.read()
            _bk_filename = f"GST_Backup_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.db"
            st.download_button(
                label="📥 Download Full Backup (.db)",
                data=_db_bytes,
                file_name=_bk_filename,
                mime="application/octet-stream",
                type="primary",
                use_container_width=True,
                help="Save this file safely. Use the Restore section below to load it on any PC."
            )
            _db_size = os.path.getsize(_db_path) / 1024
            st.caption(f"Database size: {_db_size:.1f} KB  |  File: {_db_path}")
        else:
            st.warning("No database file found yet. Run a reconciliation first.")

        st.markdown("---")

        # ── RESTORE ────────────────────────────────────────────────────────
        st.markdown("#### 📥 Restore from Backup")
        st.warning(
            "⚠️ **Restoring will REPLACE your current database** with the backup file. "
            "This cannot be undone. Make sure to export a backup first if you want to keep current data."
        )

        _restore_file = st.file_uploader(
            "Upload backup .db file",
            type=["db"],
            key="restore_db_uploader",
            help="Upload a .db file previously exported from this tool."
        )

        if _restore_file:
            _restore_bytes = _restore_file.read()
            st.info(f"File: **{_restore_file.name}** — {len(_restore_bytes)/1024:.1f} KB")

            _confirm = st.checkbox(
                "✅ I understand this will replace my current data permanently",
                key="restore_confirm"
            )
            if _confirm:
                if st.button("🔄 Restore Backup Now", type="primary", use_container_width=True, key="do_restore"):
                    try:
                        # Validate it's a SQLite file
                        if _restore_bytes[:16] != b'SQLite format 3\x00':
                            st.error("❌ Invalid file — this does not appear to be a valid SQLite database.")
                        else:
                            with open(_db_path, "wb") as _dbw:
                                _dbw.write(_restore_bytes)
                            init_db()  # run migrations on restored DB
                            # Clear session so history reloads from restored DB
                            for _k in ['last_result','df_b_clean','df_g_clean','cdnr_result',
                                       'cdnr_summary','current_recon_id','current_client_path']:
                                if _k in st.session_state:
                                    del st.session_state[_k]
                            st.session_state.app_stage = 'setup'
                            st.success("✅ Backup restored successfully! Reloading...")
                            time.sleep(1)
                            st.rerun()
                    except Exception as _re:
                        st.error(f"❌ Restore failed: {_re}")

    # B2B Download now available in 📥 Downloads Hub (Tab 3)
